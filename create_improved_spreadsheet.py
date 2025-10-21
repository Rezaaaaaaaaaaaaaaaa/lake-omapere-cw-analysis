"""
Create IMPROVED Excel spreadsheet with enhanced formatting and analysis
Project: TKIL2602 - Lake Omapere Modelling
Author: Reza Moghaddam
Date: October 2025
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# Set up paths
BASE_DIR = Path(__file__).parent
RESULTS_DIR = BASE_DIR / "Results"
RESULTS_DIR.mkdir(exist_ok=True)

print("="*80)
print("Creating IMPROVED Excel Spreadsheet - Lake Omapere CW Analysis")
print("="*80)

# Load the CSV data
csv_file = RESULTS_DIR / "CW_Coverage_by_Subcatchment.csv"
df = pd.read_csv(csv_file)

print(f"\n1. Loaded data: {len(df)} sub-catchments")

# Create enhanced Excel file
excel_file = RESULTS_DIR / "CW_Coverage_by_Subcatchment.xlsx"

# Create Excel writer
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

    # ========================================================================
    # Sheet 1: Main Data with better column names
    # ========================================================================
    print("\n2. Creating Sheet 1: Detailed Sub-catchment Analysis...")

    # Rename columns for clarity
    df_detailed = df.copy()
    df_detailed.columns = [
        'Sub-catchment ID',
        'Total Area (ha)',
        'Type 1 GW Area (ha)',
        'Type 1 GW Coverage (%)',
        'Type 2 Surface Area (ha)',
        'Type 2 Surface Coverage (%)',
        'Combined CW Area (ha)',
        'Combined Coverage (%)'
    ]

    # Add a ranking column for each type
    df_detailed['Type 1 Rank'] = df_detailed['Type 1 GW Coverage (%)'].rank(ascending=False, method='min')
    df_detailed['Type 2 Rank'] = df_detailed['Type 2 Surface Coverage (%)'].rank(ascending=False, method='min')
    df_detailed['Combined Rank'] = df_detailed['Combined Coverage (%)'].rank(ascending=False, method='min')

    # Sort by combined coverage (highest first)
    df_detailed_sorted = df_detailed.sort_values('Combined Coverage (%)', ascending=False).reset_index(drop=True)

    # Write to Excel
    df_detailed_sorted.to_excel(writer, sheet_name='Sub-catchment Analysis', index=False)

    # ========================================================================
    # Sheet 2: Type 1 (Groundwater) Focus
    # ========================================================================
    print("   Creating Sheet 2: Type 1 Groundwater CW Analysis...")

    df_type1 = df_detailed[df_detailed['Type 1 GW Coverage (%)'] > 0].copy()
    df_type1 = df_type1.sort_values('Type 1 GW Coverage (%)', ascending=False).reset_index(drop=True)
    df_type1_subset = df_type1[[
        'Sub-catchment ID',
        'Total Area (ha)',
        'Type 1 GW Area (ha)',
        'Type 1 GW Coverage (%)',
        'Type 1 Rank'
    ]]

    df_type1_subset.to_excel(writer, sheet_name='Type 1 - Groundwater', index=False)

    # ========================================================================
    # Sheet 3: Type 2 (Surface) Focus
    # ========================================================================
    print("   Creating Sheet 3: Type 2 Surface CW Analysis...")

    df_type2 = df_detailed[df_detailed['Type 2 Surface Coverage (%)'] > 0].copy()
    df_type2 = df_type2.sort_values('Type 2 Surface Coverage (%)', ascending=False).reset_index(drop=True)
    df_type2_subset = df_type2[[
        'Sub-catchment ID',
        'Total Area (ha)',
        'Type 2 Surface Area (ha)',
        'Type 2 Surface Coverage (%)',
        'Type 2 Rank'
    ]]

    df_type2_subset.to_excel(writer, sheet_name='Type 2 - Surface', index=False)

    # ========================================================================
    # Sheet 4: Summary Statistics
    # ========================================================================
    print("   Creating Sheet 4: Summary Statistics...")

    summary_data = {
        'Metric': [
            'Total Number of Sub-catchments',
            'Total Catchment Area (ha)',
            '',
            'TYPE 1 - SHALLOW GROUNDWATER (SUBSURFACE FLOWS)',
            'Total Type 1 CW Area (ha)',
            'Average Type 1 Coverage (%)',
            'Median Type 1 Coverage (%)',
            'Maximum Type 1 Coverage (%)',
            'Sub-catchments with Type 1 Coverage > 0',
            'Sub-catchments with Type 1 Coverage > 5%',
            'Sub-catchments with Type 1 Coverage > 10%',
            '',
            'TYPE 2 - TOPOGRAPHIC DEPRESSIONS (SURFACE FLOWS)',
            'Total Type 2 CW Area (ha)',
            'Average Type 2 Coverage (%)',
            'Median Type 2 Coverage (%)',
            'Maximum Type 2 Coverage (%)',
            'Sub-catchments with Type 2 Coverage > 0',
            'Sub-catchments with Type 2 Coverage > 5%',
            'Sub-catchments with Type 2 Coverage > 10%',
            '',
            'COMBINED CW COVERAGE',
            'Total Combined CW Area (ha)',
            'Average Combined Coverage (%)',
            'Median Combined Coverage (%)',
            'Maximum Combined Coverage (%)',
            'Sub-catchments with Any CW Coverage',
            'Sub-catchments with Coverage > 5%',
            'Sub-catchments with Coverage > 10%',
        ],
        'Value': [
            len(df),
            round(df['Subcatchment_Area_ha'].sum(), 2),
            '',
            '',
            round(df['Type1_GW_Area_ha'].sum(), 2),
            round(df['Type1_GW_Percent'].mean(), 2),
            round(df['Type1_GW_Percent'].median(), 2),
            round(df['Type1_GW_Percent'].max(), 2),
            (df['Type1_GW_Percent'] > 0).sum(),
            (df['Type1_GW_Percent'] > 5).sum(),
            (df['Type1_GW_Percent'] > 10).sum(),
            '',
            '',
            round(df['Type2_SW_Area_ha'].sum(), 2),
            round(df['Type2_SW_Percent'].mean(), 2),
            round(df['Type2_SW_Percent'].median(), 2),
            round(df['Type2_SW_Percent'].max(), 2),
            (df['Type2_SW_Percent'] > 0).sum(),
            (df['Type2_SW_Percent'] > 5).sum(),
            (df['Type2_SW_Percent'] > 10).sum(),
            '',
            '',
            round(df['Combined_Area_ha'].sum(), 2),
            round(df['Combined_Percent'].mean(), 2),
            round(df['Combined_Percent'].median(), 2),
            round(df['Combined_Percent'].max(), 2),
            (df['Combined_Percent'] > 0).sum(),
            (df['Combined_Percent'] > 5).sum(),
            (df['Combined_Percent'] > 10).sum(),
        ]
    }

    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Summary Statistics', index=False)

    # ========================================================================
    # Sheet 5: Top 10 Rankings
    # ========================================================================
    print("   Creating Sheet 5: Top 10 Rankings...")

    top10_type1 = df_detailed.nlargest(10, 'Type 1 GW Coverage (%)')[
        ['Sub-catchment ID', 'Total Area (ha)', 'Type 1 GW Coverage (%)']
    ].reset_index(drop=True)
    top10_type1.index = top10_type1.index + 1

    top10_type2 = df_detailed.nlargest(10, 'Type 2 Surface Coverage (%)')[
        ['Sub-catchment ID', 'Total Area (ha)', 'Type 2 Surface Coverage (%)']
    ].reset_index(drop=True)
    top10_type2.index = top10_type2.index + 1

    top10_combined = df_detailed.nlargest(10, 'Combined Coverage (%)')[
        ['Sub-catchment ID', 'Total Area (ha)', 'Combined Coverage (%)']
    ].reset_index(drop=True)
    top10_combined.index = top10_combined.index + 1

    # Write all three rankings side by side
    startrow = 0
    top10_type1.to_excel(writer, sheet_name='Top 10 Rankings', startrow=startrow, startcol=0)
    top10_type2.to_excel(writer, sheet_name='Top 10 Rankings', startrow=startrow, startcol=5)
    top10_combined.to_excel(writer, sheet_name='Top 10 Rankings', startrow=startrow, startcol=10)

print("\n3. Applying formatting to Excel file...")

# Load the workbook for formatting
wb = load_workbook(excel_file)

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========================================================================
# Format Sheet 1: Sub-catchment Analysis
# ========================================================================
ws1 = wb['Sub-catchment Analysis']

# Set column widths
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 20
ws1.column_dimensions['H'].width = 20
ws1.column_dimensions['I'].width = 13
ws1.column_dimensions['J'].width = 13
ws1.column_dimensions['K'].width = 15

# Format headers
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_thin

# Add color scale for percentage columns
ws1.conditional_formatting.add('D2:D51',
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   mid_type='num', mid_value=10, mid_color='9ECAE1',
                   end_type='num', end_value=20, end_color='08519C'))

ws1.conditional_formatting.add('F2:F51',
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   mid_type='num', mid_value=10, mid_color='A1D99B',
                   end_type='num', end_value=22, end_color='006D2C'))

ws1.conditional_formatting.add('H2:H51',
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   mid_type='num', mid_value=10, mid_color='FDAE6B',
                   end_type='num', end_value=22, end_color='D94701'))

# ========================================================================
# Format Sheet 2: Type 1 - Groundwater
# ========================================================================
ws2 = wb['Type 1 - Groundwater']

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 20
ws2.column_dimensions['E'].width = 13

for cell in ws2[1]:
    cell.fill = PatternFill(start_color='08519C', end_color='08519C', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = header_alignment
    cell.border = border_thin

ws2.conditional_formatting.add('D2:D25',
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   end_type='num', end_value=20, end_color='08519C'))

# ========================================================================
# Format Sheet 3: Type 2 - Surface
# ========================================================================
ws3 = wb['Type 2 - Surface']

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 13

for cell in ws3[1]:
    cell.fill = PatternFill(start_color='006D2C', end_color='006D2C', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = header_alignment
    cell.border = border_thin

ws3.conditional_formatting.add('D2:D30',
    ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                   end_type='num', end_value=22, end_color='006D2C'))

# ========================================================================
# Format Sheet 4: Summary Statistics
# ========================================================================
ws4 = wb['Summary Statistics']

ws4.column_dimensions['A'].width = 50
ws4.column_dimensions['B'].width = 20

for cell in ws4[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_thin

# Highlight section headers
section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
section_font = Font(bold=True, size=11)

for row in [4, 12, 21]:
    ws4[f'A{row}'].fill = section_fill
    ws4[f'A{row}'].font = section_font

# ========================================================================
# Format Sheet 5: Top 10 Rankings
# ========================================================================
ws5 = wb['Top 10 Rankings']

# Add titles
ws5['A1'] = 'Top 10 - Type 1 Groundwater'
ws5['F1'] = 'Top 10 - Type 2 Surface'
ws5['K1'] = 'Top 10 - Combined Coverage'

title_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
title_font = Font(bold=True, size=13)

for cell_ref in ['A1', 'F1', 'K1']:
    ws5[cell_ref].fill = title_fill
    ws5[cell_ref].font = title_font
    ws5[cell_ref].alignment = Alignment(horizontal='center')

# Merge title cells
ws5.merge_cells('A1:D1')
ws5.merge_cells('F1:I1')
ws5.merge_cells('K1:N1')

# Set column widths
for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N']:
    ws5.column_dimensions[col].width = 18

# Save the formatted workbook
wb.save(excel_file)

print(f"   [OK] Applied formatting and conditional coloring")

print("\n" + "="*80)
print("IMPROVED Excel Spreadsheet Created Successfully!")
print("="*80)
print(f"\nFile: {excel_file}")
print("\nSheets created:")
print("  1. Sub-catchment Analysis - Complete data with rankings")
print("  2. Type 1 - Groundwater - Focus on shallow groundwater CWs")
print("  3. Type 2 - Surface - Focus on topographic depression CWs")
print("  4. Summary Statistics - Comprehensive statistics for both types")
print("  5. Top 10 Rankings - Top performers for each CW type")
print("\nEnhancements:")
print("  - Color-coded cells (blue for Type 1, green for Type 2)")
print("  - Conditional formatting showing coverage intensity")
print("  - Rankings for easy identification of priority areas")
print("  - Detailed summary statistics")
print("  - Professional formatting with headers and borders")
print("  - Optimized column widths for readability")
