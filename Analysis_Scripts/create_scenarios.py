"""
Create scenario lookup files for Lake Omapere CW mitigation modeling.

Based on Annette's instructions, we only need single mitigation (CW) scenarios
with the existing coverage thresholds from LRFs_years.xlsx.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import shutil

# Load the template
template_path = 'Model/Lookups/Scenarios.xlsx'
output_path = 'Model/Lookups/Scenarios_LakeOmapere.xlsx'

# Copy template
shutil.copy(template_path, output_path)

# Load the workbook
wb = openpyxl.load_workbook(output_path)
ws = wb['Mitigation Classes']

print("=" * 80)
print("CREATING LAKE OMAPERE SCENARIO LOOKUP FILE")
print("=" * 80)
print()

# Clear existing mitigation scenarios (rows 3 onwards)
for row in range(3, 20):
    for col in range(1, 16):
        ws.cell(row=row, column=col).value = None

# Define CW scenarios for Lake Omapere
# Based on current LRF thresholds: <2%, 2-4%, >4%
scenarios = [
    {
        'Scenario': 1,
        'Description': 'CW <2% coverage',
        'Mitigation': 'CW',
        'ExtCode': 1,
        'BE': 0,  # Bank erosion control
        'SD': 1,  # Sediment detention
        'SR': 1,  # Sediment retention
        'TD': 1,  # Total detention
        'IF': 1,  # Infiltration
        'SG': 1,  # Shallow groundwater
        'DG': 0,  # Deep groundwater
        'Shade': 0,  # No shade
        'OrderMin': 1,
        'OrderMax': 8
    },
    {
        'Scenario': 2,
        'Description': 'CW 2-4% coverage',
        'Mitigation': 'CW',
        'ExtCode': 2,
        'BE': 0,
        'SD': 1,
        'SR': 1,
        'TD': 1,
        'IF': 1,
        'SG': 1,
        'DG': 0,
        'Shade': 0,
        'OrderMin': 1,
        'OrderMax': 8
    },
    {
        'Scenario': 3,
        'Description': 'CW >4% coverage',
        'Mitigation': 'CW',
        'ExtCode': 3,
        'BE': 0,
        'SD': 1,
        'SR': 1,
        'TD': 1,
        'IF': 1,
        'SG': 1,
        'DG': 0,
        'Shade': 0,
        'OrderMin': 1,
        'OrderMax': 8
    }
]

# Write scenarios to worksheet
for i, scen in enumerate(scenarios, start=3):
    ws.cell(row=i, column=1).value = scen['Scenario']
    ws.cell(row=i, column=2).value = scen['Description']
    ws.cell(row=i, column=3).value = scen['Mitigation']
    ws.cell(row=i, column=4).value = scen['ExtCode']
    ws.cell(row=i, column=5).value = f"=_xlfn.CONCAT(C{i},\"_\",D{i})"  # Key formula
    ws.cell(row=i, column=6).value = scen['BE']
    ws.cell(row=i, column=7).value = scen['SD']
    ws.cell(row=i, column=8).value = scen['SR']
    ws.cell(row=i, column=9).value = scen['TD']
    ws.cell(row=i, column=10).value = scen['IF']
    ws.cell(row=i, column=11).value = scen['SG']
    ws.cell(row=i, column=12).value = scen['DG']
    ws.cell(row=i, column=13).value = scen['Shade']
    ws.cell(row=i, column=14).value = scen['OrderMin']
    ws.cell(row=i, column=15).value = scen['OrderMax']

    print(f"Scenario {scen['Scenario']}: {scen['Description']}")

# Save the workbook
wb.save(output_path)
print()
print(f"[SAVED] Lake Omapere scenarios saved to:")
print(f"        {output_path}")
print()

print("=" * 80)
print("SCENARIO DETAILS")
print("=" * 80)
print()
print("The scenarios are configured for CW (Constructed Wetland) mitigation with")
print("three coverage levels matching the LRF thresholds:")
print()
print("  Scenario 1: CW <2% coverage  (ExtCode 1)")
print("  Scenario 2: CW 2-4% coverage (ExtCode 2)")
print("  Scenario 3: CW >4% coverage  (ExtCode 3)")
print()
print("All scenarios include the following processes:")
print("  - Sediment detention (SD)")
print("  - Sediment retention (SR)")
print("  - Total detention (TD)")
print("  - Infiltration (IF)")
print("  - Shallow groundwater interaction (SG)")
print()
print("Stream order range: 1-8 (all streams in catchment)")
print()
print("=" * 80)
print()
print("NEXT STEPS:")
print("  1. Update StandAloneDNZ2.py to reference this scenario file")
print("  2. Ensure LRF values in LRFs_years.xlsx match these ExtCodes")
print("  3. Run model for baseline (no mitigation)")
print("  4. Run model for wetland scenario (+0.66m)")
print()
print("=" * 80)
