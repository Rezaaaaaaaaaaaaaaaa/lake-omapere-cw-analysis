#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lake Omapere CW Mitigation Effectiveness Analysis
====================================================

Comprehensive analysis tool that processes CLUES generated loads, applies CW mitigation,
routes loads through the reach network, and generates results.

Pipeline:
1. Load CLUES spreadsheets (baseline and wetland scenarios)
2. Extract TP loads and reach identifiers
3. Read CW coverage from GIS/CSV
4. Calculate generated loads for all scenarios
5. Apply CW mitigation with LRF factors
6. Route loads through reach network with attenuation
7. Generate results files, visualizations, and summaries

Usage:
    python lake_omapere_cw_analysis.py

Author: Analysis Team
Date: October 2025
Project: TKIL2602 - Lake Omapere Modelling
"""

import os
import sys

# Set UTF-8 encoding for Windows console
if sys.platform.startswith('win'):
    try:
        import locale
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass  # Older Python versions
import csv
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import json

# Try to import optional visualization libraries
try:
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Warning: matplotlib not available. Visualizations will be skipped.")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel input may be limited.")

try:
    import geopandas as gpd
    import matplotlib.colors as mcolors
    from matplotlib.collections import LineCollection
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False
    print("Warning: geopandas not available. Mapping will be skipped.")


# ============================================================================
# CONFIGURATION
# ============================================================================

class Config:
    """Configuration for Lake Omapere analysis"""

    # File paths
    CLUES_BASELINE_PATH = "Model/InputData/CLUESloads_baseline.csv"
    CLUES_WETLAND_PATH = "Model/InputData/CLUESloads_wetland_066m.csv"
    CW_COVERAGE_XLSX = "CW_Coverage_GIS_CALCULATED.xlsx"
    REACH_NETWORK_CSV = "Model/InputData/Hydroedge2_5.csv"
    ATTENUATION_CSV = "Model/InputData/AttenCarry.csv"
    FSL_DATA_CSV = "Model/InputData/FSLData.csv"
    LRF_XLSX = "Model/Lookups/LRFs_years.xlsx"

    # Additional data files for P fractions and pathways (Annette's methodology)
    CONTAMINANT_SPLITS_XLSX = "Model/Lookups/ContaminantSplits.xlsx"
    LANDUSE_CSV = "Model/InputData/DefLUREC2_5.csv"
    HYPE_CSV = "Model/InputData/Hype.csv"

    # Shapefile paths for mapping
    RIVER_SHAPEFILE = "Shapefiles/Reference/Riverlines.shp"
    CATCHMENT_SHAPEFILE = "Shapefiles/Reference/Catchment.shp"
    LAKE_SHAPEFILE = "Shapefiles/Lake/Lake Omapere-236.34 mAMSL (+0.66m).shp"
    SUBCATCHMENTS_SHAPEFILE = "Shapefiles/Subcatchments/Subs.shp"

    # Output paths
    OUTPUT_DIR = "Results/LAKE_OMAPERE_RESULTS"
    DATA_DIR = "Results/LAKE_OMAPERE_RESULTS/Data"
    FIGURES_DIR = "Results/LAKE_OMAPERE_RESULTS/Figures"
    MAPS_DIR = "Results/LAKE_OMAPERE_RESULTS/Maps"
    SUMMARY_DIR = "Results/LAKE_OMAPERE_RESULTS/Summary"

    # CLUES column mappings (Excel columns)
    CLUES_COLUMNS = {
        'reach_id': 'NZSEGMENT',
        'overseer_load': 'AF',  # Agricultural TP (t/y)
        'sediment_p': 'T',       # Sediment P (t/y)
        'load_increment': 'BC'   # Total TP load (t/y)
    }

    # CW coverage categories and LRF factors
    # Values from CW Practitioner Guide (as requested by Fleur - Nov 3, 2025)
    # LRF = Load Reduction Factor (the fraction removed by CW treatment)
    LRF_FACTORS = {
        'low': 0.26,      # <2% coverage: 26% reduction
        'medium': 0.42,   # 2-4% coverage: 42% reduction
        'high': 0.48      # >4% coverage: 48% reduction
    }

    # Coverage thresholds (%)
    COVERAGE_THRESHOLDS = {
        'low': 2.0,
        'medium': 4.0
    }

    # P Fraction splits (from Annette's ContaminantSplits.xlsx)
    P_FRACTIONS = {
        'PartP': 0.50,  # Particulate P
        'DRP': 0.25,    # Dissolved Reactive P
        'DOP': 0.25     # Dissolved Organic P
    }

    # Pathway names (from HYPE and Bank Erosion)
    PATHWAYS = ['SR', 'SD', 'TD', 'IF', 'SG', 'DG', 'BE']

    # LRF pathway mapping (from LRFs_years.xlsx CW sheet)
    # Maps coverage category to pathway-specific LRFs
    LRF_PATHWAY_MAPPING = {
        'low': {'ExtCode': 1, 'Extent': '<2'},       # <2% coverage
        'medium': {'ExtCode': 2, 'Extent': '2 to 4'}, # 2-4% coverage
        'high': {'ExtCode': 3, 'Extent': '>4'}        # >4% coverage
    }

    # Lake Omapere specific reaches (50 reaches to analyze)
    LAKE_REACHES = None  # Will be loaded from data

    # Attenuation parameters
    DEFAULT_ATTENUATION = 0.90  # If PstreamCarry not available
    CLAY_THRESHOLD = 50.0       # % clay to flag as HighClay


# ============================================================================
# SECTION 1: DATA LOADING
# ============================================================================

class DataLoader:
    """Load and parse input data from various sources"""

    @staticmethod
    def load_clues_excel(filepath, scenario_name="baseline"):
        """
        Load CLUES data and extract TP loads.

        Args:
            filepath: Path to CLUES file (.csv, .xlsb, or .xlsx)
            scenario_name: 'baseline' or 'wetland'

        Returns:
            DataFrame with CLUES columns including: NZSEGMENT, TPAgGen, soilP, TPGen
        """
        print(f"\nLoading CLUES {scenario_name} from: {filepath}")

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"CLUES file not found: {filepath}")

        try:
            # Handle CSV files
            if filepath.endswith('.csv'):
                df = pd.read_csv(filepath)
                print(f"  Loaded {len(df)} rows from CSV")

                # Verify required columns exist
                required_cols = ['NZSEGMENT', 'TPAgGen', 'soilP', 'TPGen']
                missing = [col for col in required_cols if col not in df.columns]
                if missing:
                    print(f"  WARNING: Missing columns: {missing}")
                    print(f"  Available columns: {df.columns.tolist()}")
                else:
                    print(f"  [OK] All required TP columns present")

                return df

            # Handle Excel files
            elif filepath.endswith('.xlsx'):
                df = pd.read_excel(filepath, sheet_name=0, header=0)
                print(f"  Loaded {len(df)} rows from .xlsx")
                return df

            # Handle xlsb files (may not work without pyxlsb)
            elif filepath.endswith('.xlsb'):
                print(f"  Note: .xlsb file detected. Consider converting to .xlsx or .csv")
                df = pd.read_excel(filepath, sheet_name=0, header=0)
                print(f"  Loaded {len(df)} rows from .xlsb")
                return df

        except Exception as e:
            print(f"  [ERROR] Error loading CLUES file: {e}")
            print(f"  Ensure file format is .csv, .xlsx, or .xlsb")
            raise

    @staticmethod
    def load_cw_coverage(filepath):
        """
        Load CW coverage percentages by reach.

        Args:
            filepath: Path to Excel/CSV with CW coverage data

        Returns:
            DataFrame with reach_id and cw_coverage_percent columns
        """
        print(f"\nLoading CW coverage from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: CW coverage file not found, using zeros")
            return pd.DataFrame({'reach_id': [], 'cw_coverage_percent': []})

        # Read Excel or CSV based on file extension
        if filepath.endswith('.xlsx') or filepath.endswith('.xlsb'):
            df = pd.read_excel(filepath)
            # Map GIS-calculated columns to expected format
            # nzsegment -> reach_id, Combined_Percent -> cw_coverage_percent
            if 'nzsegment' in df.columns and 'Combined_Percent' in df.columns:
                df = df.rename(columns={
                    'nzsegment': 'reach_id',
                    'Combined_Percent': 'cw_coverage_percent'
                })
                # Keep additional useful columns if available
                keep_cols = ['reach_id', 'cw_coverage_percent']
                optional_cols = ['Combined_CW_Area_ha', 'Land_Area_ha_GIS',
                                'Type1_GW_Percent', 'Type2_SW_Percent']
                for col in optional_cols:
                    if col in df.columns:
                        keep_cols.append(col)
                df = df[keep_cols]
        else:
            df = pd.read_csv(filepath)

        print(f"  Loaded {len(df)} reaches with CW data")
        print(f"  CW coverage range: {df['cw_coverage_percent'].min():.2f}% to {df['cw_coverage_percent'].max():.2f}%")
        return df

    @staticmethod
    def load_reach_network(filepath):
        """
        Load reach network connectivity.

        Args:
            filepath: Path to CSV with FROM_NODE, TO_NODE, and HYDSEQ columns

        Returns:
            DataFrame with network connectivity
        """
        print(f"\nLoading reach network from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: Network file not found")
            return None

        df = pd.read_csv(filepath)
        print(f"  Loaded network with {len(df)} reaches")
        return df

    @staticmethod
    def load_attenuation_factors(filepath):
        """
        Load attenuation (PstreamCarry) factors for reaches.

        Args:
            filepath: Path to CSV with reach_id and PstreamCarry columns

        Returns:
            DataFrame with attenuation factors
        """
        print(f"\nLoading attenuation factors from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: Attenuation file not found, using defaults")
            return None

        df = pd.read_csv(filepath)
        print(f"  Loaded attenuation factors for {len(df)} reaches")
        return df

    @staticmethod
    def load_clay_data(filepath='Model/InputData/FSLData.csv'):
        """
        Load clay soil data (Clayey boolean flag from FSL).

        Args:
            filepath: Path to FSLData.csv with NZSEGMENT and Clayey columns

        Returns:
            DataFrame with reach_id and clay_percent columns
            (Clayey=0 → 0%, Clayey=1 → 100%)
        """
        print(f"\nLoading clay soil data from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: Clay data file not found, assuming no clayey soils")
            return None

        try:
            df = pd.read_csv(filepath, encoding="Latin1")
            if 'NZSEGMENT' in df.columns and 'Clayey' in df.columns:
                clay_df = df[['NZSEGMENT', 'Clayey']].copy()
                clay_df.columns = ['reach_id', 'clayey_soil']
                clay_df['clayey_soil'] = clay_df['clayey_soil'].fillna(0).astype(int)

                # Convert boolean to percentage: 0 → 0%, 1 → 100%
                clay_df['clay_percent'] = clay_df['clayey_soil'] * 100.0

                print(f"  Loaded clay data for {len(clay_df)} reaches")
                print(f"  Reaches with clayey soil (Clayey=1): {(clay_df['clayey_soil']==1).sum()}")

                return clay_df[['reach_id', 'clay_percent']]
            else:
                print(f"  Warning: Required columns not found in clay data")
                return None
        except Exception as e:
            print(f"  Warning: Error loading clay data: {e}")
            return None

    @staticmethod
    def load_p_fractions(filepath):
        """
        Load P fraction splits from ContaminantSplits.xlsx.

        Args:
            filepath: Path to ContaminantSplits.xlsx

        Returns:
            Dictionary with P fraction splits (PartP, DRP, DOP)
        """
        print(f"\nLoading P fraction splits from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: P fractions file not found, using defaults")
            return Config.P_FRACTIONS

        try:
            df = pd.read_excel(filepath, sheet_name='P')
            fractions = {}
            for _, row in df.iterrows():
                fractions[row['Form']] = row['Fraction']

            print(f"  Loaded P fractions: {fractions}")
            return fractions
        except Exception as e:
            print(f"  Warning: Error loading P fractions: {e}, using defaults")
            return Config.P_FRACTIONS

    @staticmethod
    def load_landuse(filepath):
        """
        Load land use percentages from DefLUREC2_5.csv.

        Args:
            filepath: Path to DefLUREC2_5.csv

        Returns:
            DataFrame with reach_id and land use percentage columns
        """
        print(f"\nLoading land use data from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: Land use file not found")
            return None

        try:
            df = pd.read_csv(filepath)
            print(f"  Loaded land use for {len(df)} reaches")

            # Rename NZSEGMENT to reach_id for consistency
            if 'NZSEGMENT' in df.columns:
                df = df.rename(columns={'NZSEGMENT': 'reach_id'})

            # Key columns to keep
            lu_cols = ['reach_id', 'DAIRY', 'SBINTEN', 'SBHILL', 'SBHIGH', 'DEER',
                      'OTHER_ANIM', 'MAIZE', 'PLANT_FOR', 'NAT_FOR', 'SCRUB',
                      'URBAN', 'Pasture', 'HortCrop', 'All']

            # Keep only columns that exist
            keep_cols = [col for col in lu_cols if col in df.columns]
            df = df[keep_cols]

            print(f"  Retained {len(keep_cols)} land use columns")
            return df
        except Exception as e:
            print(f"  Warning: Error loading land use: {e}")
            return None

    @staticmethod
    def load_hype_pathways(filepath):
        """
        Load HYPE pathway percentages from Hype.csv.

        Args:
            filepath: Path to Hype.csv

        Returns:
            DataFrame with reach_id and pathway percentages (SR, TD, IF, SG, DG)
        """
        print(f"\nLoading HYPE pathway data from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: HYPE file not found")
            return None

        try:
            df = pd.read_csv(filepath)
            print(f"  Loaded HYPE pathways for {len(df)} reaches")

            # Rename NZSEGMENT to reach_id
            if 'NZSEGMENT' in df.columns:
                df = df.rename(columns={'NZSEGMENT': 'reach_id'})

            # Keep pathway columns
            pathway_cols = ['reach_id', 'HYDSEQ', 'SR', 'TD', 'IF', 'SG', 'DG']
            keep_cols = [col for col in pathway_cols if col in df.columns]
            df = df[keep_cols]

            # Replace nulls with default values (equal split: 25% each for SR, IF, SG, DG; 0% for TD)
            df['SR'] = df['SR'].fillna(25.0)
            df['IF'] = df['IF'].fillna(25.0)
            df['SG'] = df['SG'].fillna(25.0)
            df['DG'] = df['DG'].fillna(25.0)
            df['TD'] = df['TD'].fillna(0.0)

            print(f"  HYPE pathway ranges:")
            for col in ['SR', 'TD', 'IF', 'SG', 'DG']:
                if col in df.columns:
                    print(f"    {col}: {df[col].min():.1f}% - {df[col].max():.1f}%")

            return df
        except Exception as e:
            print(f"  Warning: Error loading HYPE pathways: {e}")
            return None

    @staticmethod
    def load_pathway_lrfs(filepath):
        """
        Load pathway-specific LRFs from LRFs_years.xlsx.

        Args:
            filepath: Path to LRFs_years.xlsx

        Returns:
            DataFrame with pathway-specific LRF values for CW mitigation
        """
        print(f"\nLoading pathway-specific LRFs from: {filepath}")

        if not os.path.exists(filepath):
            print(f"  Warning: LRF file not found")
            return None

        try:
            df = pd.read_excel(filepath, sheet_name='CW')
            print(f"  Loaded {len(df)} LRF entries for CW mitigation")

            # Show available columns for P fractions
            p_cols = [col for col in df.columns if any(x in col for x in ['PartP', 'DRP', 'DOP'])]
            print(f"  P fraction LRF columns available: {p_cols}")

            return df
        except Exception as e:
            print(f"  Warning: Error loading pathway LRFs: {e}")
            return None


# ============================================================================
# SECTION 2: GENERATED LOADS CALCULATION
# ============================================================================

class GeneratedLoadsCalculator:
    """Calculate TP loads generated at reach level"""

    @staticmethod
    def extract_load_components(clues_df, reach_col='NZSEGMENT'):
        """
        Extract TP load components from CLUES data.

        Extracts:
        - TPAgGen: Agricultural TP (OVERSEER Load)
        - soilP: Sediment phosphorus
        - TPGen: Non-pastoral TP (calculated)

        Args:
            clues_df: CLUES DataFrame
            reach_col: Column name for reach ID

        Returns:
            DataFrame with load components
        """
        print(f"\nExtracting load components from CLUES data...")

        # Map column names
        results = pd.DataFrame({
            'reach_id': clues_df[reach_col],
        })

        # Extract load components
        # From Annette's instructions: Total TP = TPAgGen + soilP + TPGen
        # where TPGen is the non-pastoral component

        # Agricultural/pastoral load
        if 'OVERSEER Load (t/y)' in clues_df.columns:
            results['ag_load'] = clues_df['OVERSEER Load (t/y)']
        elif 'TPAgGen' in clues_df.columns:
            results['ag_load'] = clues_df['TPAgGen']
        else:
            results['ag_load'] = 0
            print("  WARNING: No agricultural load column found")

        # Sediment phosphorus
        if 'P_Sed' in clues_df.columns:
            results['sediment_p'] = clues_df['P_Sed']
        elif 'soilP' in clues_df.columns:
            results['sediment_p'] = clues_df['soilP']
        else:
            results['sediment_p'] = 0
            print("  WARNING: No sediment P column found")

        # Non-pastoral load
        if 'TPGen' in clues_df.columns:
            results['non_pastoral_tp'] = clues_df['TPGen']
        else:
            results['non_pastoral_tp'] = 0
            print("  WARNING: No TPGen column found")

        # TOTAL LOAD = Agricultural + Sediment + Non-pastoral
        # This is the CORRECT calculation per Annette's instructions
        results['total_load'] = (results['ag_load'] +
                                results['sediment_p'] +
                                results['non_pastoral_tp'])

        print(f"  Extracted components for {len(results)} reaches")
        print(f"  Mean total load: {results['total_load'].mean():.4f} t/y")

        return results

    @staticmethod
    def calculate_generated_loads(baseline_loads, wetland_loads):
        """
        Calculate generated loads for baseline and wetland scenarios.

        Args:
            baseline_loads: DataFrame with baseline TP loads
            wetland_loads: DataFrame with wetland scenario TP loads

        Returns:
            DataFrame with baseline and wetland generated loads
        """
        print(f"\nCalculating generated loads...")

        # Merge baseline and wetland data
        results = baseline_loads[['reach_id', 'total_load']].copy()
        results.rename(columns={'total_load': 'generated_baseline'},
                      inplace=True)

        # Add wetland scenario
        wetland_subset = wetland_loads[['reach_id', 'total_load']].copy()
        wetland_subset.rename(columns={'total_load': 'generated_wetland'},
                             inplace=True)

        results = results.merge(wetland_subset, on='reach_id', how='left')

        # Fill missing values
        results['generated_wetland'].fillna(results['generated_baseline'],
                                           inplace=True)

        print(f"  Baseline mean: {results['generated_baseline'].mean():.4f} t/y")
        print(f"  Wetland mean:  {results['generated_wetland'].mean():.4f} t/y")

        return results


# ============================================================================
# SECTION 2B: P FRACTION AND PATHWAY SPLITTING (Annette's Methodology)
# ============================================================================

class PFractionPathwayCalculator:
    """Split TP loads into P fractions and HYPE pathways following Annette's model"""

    @staticmethod
    def split_tp_into_fractions(tp_loads, p_fractions, include_sources=True):
        """
        Split total TP into P fractions (PartP, DRP, DOP).

        Following Annette's Splitter.py methodology:
        - PartP (Particulate P): 50% of TP
        - DRP (Dissolved Reactive P): 25% of TP
        - DOP (Dissolved Organic P): 25% of TP

        Args:
            tp_loads: DataFrame with TP load columns
            p_fractions: Dict with fraction splits {'PartP': 0.50, 'DRP': 0.25, 'DOP': 0.25}
            include_sources: If True, split by source (Ag, nonAg, ps)

        Returns:
            DataFrame with P fraction columns added
        """
        print("\nSplitting TP into P fractions...")

        results = tp_loads.copy()

        # Get column names for different scenarios
        tp_columns = [col for col in tp_loads.columns if 'generated' in col or 'total_load' in col]

        for tp_col in tp_columns:
            for fract_name, fract_pct in p_fractions.items():
                new_col = f"{tp_col}_{fract_name}"
                results[new_col] = results[tp_col] * fract_pct
                print(f"  Created {new_col}: {fract_pct*100}% of {tp_col}")

        print(f"  Split complete for {len(results)} reaches")
        return results

    @staticmethod
    def split_by_hype_pathways(loads_df, hype_df, p_fractions):
        """
        Split P fractions into HYPE pathways (SR, SD, TD, IF, SG, DG).

        Following Annette's methodology:
        - Surface Runoff (SR): Variable %
        - Surface Drainage (SD): Half of SR where tile drains present (TD > 0)
        - Tile Drainage (TD): Variable %
        - Interflow (IF): Variable %
        - Shallow Groundwater (SG): Variable %
        - Deep Groundwater (DG): Variable %

        Bank Erosion (BE) is split based on sediment erosion percentage.

        Args:
            loads_df: DataFrame with P fraction columns
            hype_df: DataFrame with HYPE pathway percentages
            p_fractions: Dict with P fraction names

        Returns:
            DataFrame with pathway-specific P fraction columns
        """
        print("\nSplitting P fractions into HYPE pathways...")

        # Merge HYPE data
        results = loads_df.merge(hype_df, on='reach_id', how='left')

        # Fill missing HYPE values with defaults
        results['SR'] = results['SR'].fillna(25.0)
        results['TD'] = results['TD'].fillna(0.0)
        results['IF'] = results['IF'].fillna(25.0)
        results['SG'] = results['SG'].fillna(25.0)
        results['DG'] = results['DG'].fillna(25.0)

        # Calculate SD (Surface Drainage) - half of SR where TD > 0
        results['SD'] = 0.0
        has_td = results['TD'] > 0
        results.loc[has_td, 'SD'] = results.loc[has_td, 'SR'] / 2.0
        results.loc[has_td, 'SR'] = results.loc[has_td, 'SR'] / 2.0

        print(f"  Reaches with tile drainage (SD calculated): {has_td.sum()}")

        # Split each P fraction by pathway
        pathways = ['SR', 'SD', 'TD', 'IF', 'SG', 'DG']

        # Find REMAINDER columns ONLY (not BE columns)
        # Only remainder columns should be distributed across HYPE pathways
        p_fract_cols = []
        for col in loads_df.columns:
            if any(fract in col for fract in p_fractions.keys()) and '_remainder' in col:
                p_fract_cols.append(col)

        print(f"  Splitting {len(p_fract_cols)} P fraction remainder columns across {len(pathways)} pathways")

        for p_col in p_fract_cols:
            for pathway in pathways:
                # Calculate pathway-specific load
                pathway_col = f"{p_col}_{pathway}"
                results[pathway_col] = results[p_col] * (results[pathway] / 100.0)

        print(f"  Created {len(p_fract_cols) * len(pathways)} pathway-specific columns")

        return results

    @staticmethod
    def calculate_bank_erosion_split(loads_df, sediment_pct=0.6):
        """
        Calculate Bank Erosion (BE) component.

        Following Annette's methodology:
        - For N and P: BE is based on sediment bank erosion percentage from NZSYE
        - Default assumption: 60% of load comes from bank erosion
        - Remainder goes through HYPE pathways

        Args:
            loads_df: DataFrame with P loads
            sediment_pct: Bank erosion percentage (default 0.6 = 60%)

        Returns:
            DataFrame with BE columns added
        """
        print("\nCalculating Bank Erosion (BE) split...")

        results = loads_df.copy()

        # Find P fraction columns to split
        p_fract_cols = [col for col in loads_df.columns
                       if any(fract in col for fract in Config.P_FRACTIONS.keys())
                       and not any(path in col for path in Config.PATHWAYS)]

        for p_col in p_fract_cols:
            # BE load
            be_col = f"{p_col}_BE"
            results[be_col] = results[p_col] * sediment_pct

            # Remainder for HYPE pathways
            remainder_col = f"{p_col}_remainder"
            results[remainder_col] = results[p_col] * (1 - sediment_pct)

        print(f"  Bank erosion split applied to {len(p_fract_cols)} P fraction columns")
        print(f"  BE percentage: {sediment_pct*100}%")

        return results


# ============================================================================
# SECTION 3: CW MITIGATION APPLICATION
# ============================================================================

class CWMitigationCalculator:
    """Apply CW mitigation with load reduction factors"""

    @staticmethod
    def categorize_coverage(coverage_percent):
        """
        Categorize CW coverage level.

        Args:
            coverage_percent: CW coverage as percentage

        Returns:
            Category: 'high', 'medium', 'low', 'none'
        """
        if coverage_percent >= Config.COVERAGE_THRESHOLDS['medium']:
            return 'high'
        elif coverage_percent >= Config.COVERAGE_THRESHOLDS['low']:
            return 'medium'
        elif coverage_percent > 0:
            return 'low'
        else:
            return 'none'

    @staticmethod
    def apply_cw_mitigation(generated_loads, cw_coverage_df,
                           clay_analysis_df=None):
        """
        Apply CW mitigation to wetland loads using LRFs.

        Formula:
        CW_Reduction = Generated_Wetland × LRF
        Mitigated_Load = Generated_Wetland - CW_Reduction

        Where LRF (Load Reduction Factor) from CW Practitioner Guide:
        - <2% coverage: LRF = 0.26 (26% reduction)
        - 2-4% coverage: LRF = 0.42 (42% reduction)
        - >4% coverage: LRF = 0.48 (48% reduction)

        Note: LRF represents the fraction of load REMOVED by CW treatment.
        Higher CW coverage = higher LRF = higher removal percentage.

        Args:
            generated_loads: DataFrame with generated loads
            cw_coverage_df: DataFrame with CW coverage percentages
            clay_analysis_df: Optional DataFrame with clay content

        Returns:
            DataFrame with CW mitigation applied
        """
        print(f"\nApplying CW mitigation...")

        results = generated_loads.copy()

        # Merge with CW coverage
        if cw_coverage_df is not None and len(cw_coverage_df) > 0:
            # Adjust column names as needed
            if 'CW_Coverage_Percent' not in cw_coverage_df.columns:
                # Try common alternatives
                coverage_col = [col for col in cw_coverage_df.columns
                              if 'coverage' in col.lower() or '%' in col.lower()]
                if coverage_col:
                    cw_coverage_df = cw_coverage_df.rename(
                        columns={coverage_col[0]: 'CW_Coverage_Percent'})

            results = results.merge(
                cw_coverage_df[['reach_id', 'CW_Coverage_Percent']],
                on='reach_id', how='left'
            )
            results['CW_Coverage_Percent'].fillna(0, inplace=True)
        else:
            results['CW_Coverage_Percent'] = 0

        # Add clay analysis if provided
        if clay_analysis_df is not None:
            results = results.merge(
                clay_analysis_df[['reach_id', 'clay_percent']],
                on='reach_id', how='left'
            )
            results['clay_percent'].fillna(0, inplace=True)
            results['HighClay'] = (results['clay_percent'] >
                                  Config.CLAY_THRESHOLD)
        else:
            results['clay_percent'] = 0
            results['HighClay'] = False

        # Categorize coverage
        results['coverage_category'] = results['CW_Coverage_Percent'].apply(
            CWMitigationCalculator.categorize_coverage
        )

        # Apply LRF factors
        lrf_map = {
            'high': Config.LRF_FACTORS['high'],
            'medium': Config.LRF_FACTORS['medium'],
            'low': Config.LRF_FACTORS['low'],
            'none': 0.0  # No CW = 0% reduction
        }

        results['lrf_factor'] = results['coverage_category'].map(lrf_map)

        # CRITICAL RULE: Override LRF to 0 for clayey soils (CWs don't work in high clay)
        # Clayey soils (clay > 50%) prevent CW effectiveness
        if 'clay_percent' in results.columns:
            # If using continuous clay percent data
            clayey_mask = results['clay_percent'] > Config.CLAY_THRESHOLD
            results.loc[clayey_mask, 'lrf_factor'] = 0.0
            clay_override_count = clayey_mask.sum()
            if clay_override_count > 0:
                print(f"  [CLAY RULE] Set LRF=0 for {clay_override_count} reaches with clayey soils")
        elif 'HighClay' in results.columns:
            # If using boolean flag
            results.loc[results['HighClay'] == True, 'lrf_factor'] = 0.0
            clay_override_count = (results['HighClay'] == True).sum()
            if clay_override_count > 0:
                print(f"  [CLAY RULE] Set LRF=0 for {clay_override_count} reaches with clayey soils")

        # Calculate CW reduction and mitigated load
        # LRF = Load Reduction Factor (fraction removed)
        # Formula: Reduction = Wetland_Load × LRF
        #          Mitigated_Load = Wetland_Load × (1 - LRF)
        results['cw_reduction'] = (results['generated_wetland'] *
                                   results['lrf_factor'])

        results['generated_cw'] = (results['generated_wetland'] -
                                   results['cw_reduction'])

        # Summary statistics
        total_coverage = (cw_coverage_df['CW_Coverage_Percent'].sum()
                         if cw_coverage_df is not None else 0)
        total_reduction = results['cw_reduction'].sum()

        print(f"  Total CW coverage: {total_coverage:.2f}%")
        print(f"  Total CW reduction: {total_reduction:.4f} t/y")
        print(f"  Coverage categories:")
        print(f"    High (>4%): {(results['coverage_category']=='high').sum()}")
        print(f"    Medium (2-4%): {(results['coverage_category']=='medium').sum()}")
        print(f"    Low (<2%): {(results['coverage_category']=='low').sum()}")
        print(f"    None: {(results['coverage_category']=='none').sum()}")

        return results

    @staticmethod
    def apply_pathway_specific_mitigation(results_df, pathway_lrfs_df, p_fractions):
        """
        Apply pathway-specific CW mitigation with different LRFs for each P fraction.

        PHASE 2 RULES (from Fleur's email):
        1. Bank Erosion (BE): NOT mitigated by CWs (all fractions pass through)
        2. Deep Groundwater (DG): Dissolved P (DRP/DOP) NOT mitigated (bypasses CW)
        3. All other pathways (SR, SD, TD, IF, SG): Apply pathway-specific LRFs
           - PartP: 100% reduction (LRF = 0% remaining)
           - DRP/DOP: Variable by coverage (23%/42%/48% remaining)

        Args:
            results_df: DataFrame with pathway-specific loads
            pathway_lrfs_df: DataFrame with LRF values from LRFs_years.xlsx
            p_fractions: Dict with P fraction names

        Returns:
            DataFrame with CW-mitigated pathway loads
        """
        print("\n" + "="*80)
        print("PHASE 2: Applying Pathway-Specific CW Mitigation")
        print("="*80)

        results = results_df.copy()

        # Build LRF lookup table from median values
        # CORRECTED MAPPING: The file has ExtCode backwards!
        # ExtCode 1 has 23% remaining (77% removal) = BEST = should be HIGH coverage
        # ExtCode 3 has 48% remaining (52% removal) = WORST = should be LOW coverage
        # ExtCode: 1 = >4% (HIGH - most removal), 2 = 2-4% (MEDIUM), 3 = <2% (LOW - least removal)
        lrf_lookup = {}
        for _, row in pathway_lrfs_df.iterrows():
            if row['ExtCode'] == 1:
                coverage_cat = 'high'  # SWAPPED: 1 -> high (best removal)
            elif row['ExtCode'] == 2:
                coverage_cat = 'medium'
            else:  # ExtCode == 3
                coverage_cat = 'low'  # SWAPPED: 3 -> low (worst removal)

            # Use median values (% remaining after CW treatment)
            # Note: These are % REMAINING, so lower = more removal
            lrf_lookup[(coverage_cat, 'PartP')] = row['PartPmed'] / 100.0  # Convert % to fraction
            lrf_lookup[(coverage_cat, 'DRP')] = row['DRPmed'] / 100.0

            # Handle DOP column name variation
            if 'DOMed' in pathway_lrfs_df.columns:
                lrf_lookup[(coverage_cat, 'DOP')] = row['DOMed'] / 100.0
            elif 'DOPmed' in pathway_lrfs_df.columns:
                lrf_lookup[(coverage_cat, 'DOP')] = row['DOPmed'] / 100.0
            else:
                lrf_lookup[(coverage_cat, 'DOP')] = row['DRPmed'] / 100.0  # Fallback to DRP

        print("\nLRF Lookup Table (% remaining after CW treatment):")
        print("  Low coverage (<2%):")
        print(f"    PartP: {lrf_lookup[('low', 'PartP')]*100:.0f}% remaining (= {(1-lrf_lookup[('low', 'PartP')])*100:.0f}% reduction)")
        print(f"    DRP:   {lrf_lookup[('low', 'DRP')]*100:.0f}% remaining (= {(1-lrf_lookup[('low', 'DRP')])*100:.0f}% reduction)")
        print(f"    DOP:   {lrf_lookup[('low', 'DOP')]*100:.0f}% remaining (= {(1-lrf_lookup[('low', 'DOP')])*100:.0f}% reduction)")
        print("  Medium coverage (2-4%):")
        print(f"    PartP: {lrf_lookup[('medium', 'PartP')]*100:.0f}% remaining")
        print(f"    DRP:   {lrf_lookup[('medium', 'DRP')]*100:.0f}% remaining")
        print(f"    DOP:   {lrf_lookup[('medium', 'DOP')]*100:.0f}% remaining")
        print("  High coverage (>4%):")
        print(f"    PartP: {lrf_lookup[('high', 'PartP')]*100:.0f}% remaining")
        print(f"    DRP:   {lrf_lookup[('high', 'DRP')]*100:.0f}% remaining")
        print(f"    DOP:   {lrf_lookup[('high', 'DOP')]*100:.0f}% remaining")

        # Mitigated pathways (NOT Bank Erosion, NOT Deep GW)
        mitigated_pathways = ['SR', 'SD', 'TD', 'IF', 'SG']

        # For each P fraction
        for fraction in p_fractions.keys():
            print(f"\nProcessing {fraction}...")

            # Find wetland REMAINDER pathway columns for this fraction
            # These are columns like "generated_wetland_PartP_remainder_SR"
            wetland_pathway_cols = [col for col in results.columns
                                   if 'generated_wetland' in col
                                   and fraction in col
                                   and '_remainder_' in col
                                   and any(path in col for path in mitigated_pathways)]

            print(f"  Found {len(wetland_pathway_cols)} wetland remainder pathway columns to mitigate")

            # Apply CW mitigation to each pathway
            for wetland_col in wetland_pathway_cols:
                # Determine CW column name
                cw_col = wetland_col.replace('generated_wetland', 'generated_cw')

                # Initialize CW column with wetland values (no mitigation by default)
                results[cw_col] = results[wetland_col].copy()

                # Apply mitigation only to reaches with CW coverage > 0
                has_cw = results['CW_Coverage_Percent'] > 0

                if has_cw.sum() > 0:
                    # For each coverage category
                    for coverage_cat in ['low', 'medium', 'high']:
                        # Get reaches in this category with CW
                        mask = (results['coverage_category'] == coverage_cat) & has_cw

                        if mask.sum() > 0:
                            # Get LRF for this fraction and coverage
                            lrf = lrf_lookup.get((coverage_cat, fraction), 0.0)

                            # Apply LRF (lrf = fraction REMAINING after treatment)
                            # CW load = wetland load × lrf
                            results.loc[mask, cw_col] = results.loc[mask, wetland_col] * lrf

            # Handle Bank Erosion (BE) - NO MITIGATION
            be_col_wetland = f'generated_wetland_{fraction}_BE'
            be_col_cw = f'generated_cw_{fraction}_BE'

            if be_col_wetland in results.columns:
                # BE passes through unchanged
                results[be_col_cw] = results[be_col_wetland].copy()
                print(f"  Bank Erosion ({fraction}): NO mitigation (passes through)")

            # Handle Deep Groundwater (DG) - Dissolved P NOT mitigated
            dg_col_wetland = f'generated_wetland_{fraction}_remainder_DG'
            dg_col_cw = f'generated_cw_{fraction}_remainder_DG'

            if dg_col_wetland in results.columns:
                if fraction in ['DRP', 'DOP']:
                    # Dissolved P in DG bypasses CW (no mitigation)
                    results[dg_col_cw] = results[dg_col_wetland].copy()
                    print(f"  Deep Groundwater ({fraction}): NO mitigation (bypasses CW)")
                else:
                    # PartP in DG (should already be 0, but apply LRF anyway for consistency)
                    # Apply mitigation like other pathways
                    has_cw = results['CW_Coverage_Percent'] > 0
                    if has_cw.sum() > 0:
                        for coverage_cat in ['low', 'medium', 'high']:
                            mask = (results['coverage_category'] == coverage_cat) & has_cw
                            if mask.sum() > 0:
                                lrf = lrf_lookup.get((coverage_cat, fraction), 0.0)
                                results.loc[mask, dg_col_cw] = results.loc[mask, dg_col_wetland] * lrf

        # Recalculate total CW loads by summing pathway loads + BE loads
        print("\nRecalculating total CW loads from pathway-specific loads...")

        for fraction in p_fractions.keys():
            # Find all CW remainder pathway columns for this fraction
            cw_pathway_cols = [col for col in results.columns
                              if 'generated_cw' in col
                              and fraction in col
                              and '_remainder_' in col]

            # Add BE column
            be_col = f'generated_cw_{fraction}_BE'

            if cw_pathway_cols and be_col in results.columns:
                # Sum all remainder pathways + BE to get total P fraction load
                total_col = f'generated_cw_{fraction}'
                results[total_col] = results[be_col] + results[cw_pathway_cols].sum(axis=1)
                print(f"  {fraction}: Summed {len(cw_pathway_cols)} pathway columns + BE")
            elif cw_pathway_cols:
                # No BE column, just sum pathways
                total_col = f'generated_cw_{fraction}'
                results[total_col] = results[cw_pathway_cols].sum(axis=1)
                print(f"  {fraction}: Summed {len(cw_pathway_cols)} pathway columns (no BE)")
            elif be_col in results.columns:
                # No pathway columns, just use BE
                total_col = f'generated_cw_{fraction}'
                results[total_col] = results[be_col]
                print(f"  {fraction}: Using BE only (no pathways)")

        # Recalculate total TP for CW scenario
        generated_cw_total = 0
        for fraction in p_fractions.keys():
            total_col = f'generated_cw_{fraction}'
            if total_col in results.columns:
                generated_cw_total += results[total_col]

        results['generated_cw'] = generated_cw_total

        # Recalculate CW reduction
        results['cw_reduction'] = results['generated_wetland'] - results['generated_cw']

        # Statistics
        total_wetland = results['generated_wetland'].sum()
        total_cw = results['generated_cw'].sum()
        total_reduction = results['cw_reduction'].sum()
        pct_reduction = (total_reduction / total_wetland * 100) if total_wetland > 0 else 0

        print("\n" + "="*80)
        print("PHASE 2 MITIGATION RESULTS:")
        print("="*80)
        print(f"Total wetland load:  {total_wetland:.2f} t/y")
        print(f"Total CW load:       {total_cw:.2f} t/y")
        print(f"Total CW reduction:  {total_reduction:.2f} t/y ({pct_reduction:.1f}%)")
        print(f"Reaches with CW:     {(results['CW_Coverage_Percent'] > 0).sum()}")
        print("="*80)

        return results


# ============================================================================
# SECTION 4: NETWORK ROUTING
# ============================================================================

class NetworkRouter:
    """Route loads through reach network with attenuation"""

    @staticmethod
    def build_network_structure(reach_network_df):
        """
        Build network structure for routing.

        Args:
            reach_network_df: DataFrame with FROM_NODE, TO_NODE, HYDSEQ

        Returns:
            Dictionary with network connectivity
        """
        print(f"\nBuilding network structure...")

        network = {
            'reaches': set(reach_network_df['NZSEGMENT'].unique()),
            'upstream': defaultdict(list),  # reach -> list of upstream reaches
            'downstream': defaultdict(list),  # reach -> downstream reach
            'hydseq': {}  # reach -> hydseq value
        }

        # Build connectivity
        for _, row in reach_network_df.iterrows():
            from_node = row.get('FROM_NODE', row.get('from_node'))
            to_node = row.get('TO_NODE', row.get('to_node'))
            reach_id = row.get('NZSEGMENT', row.get('reach_id'))
            hydseq = row.get('HYDSEQ', 0)

            # Find which reach corresponds to each node
            # (This is simplified - actual implementation may need mapping)

            network['hydseq'][reach_id] = hydseq

        print(f"  Network has {len(network['reaches'])} reaches")
        return network

    @staticmethod
    def route_loads(loads_df, reach_network_df, attenuation_df=None):
        """
        Route loads through reach network with attenuation.

        Algorithm:
        For each reach (in HYDSEQ order):
            Routed[i] = Generated[i] + Σ(Upstream_Routed[j] × Attenuation[i])

        Args:
            loads_df: DataFrame with generated loads
            reach_network_df: DataFrame with network connectivity
            attenuation_df: DataFrame with attenuation factors

        Returns:
            DataFrame with routed loads
        """
        print(f"\nRouting loads through network...")

        results = loads_df.copy()

        # Merge attenuation factors
        if attenuation_df is not None and len(attenuation_df) > 0:
            results = results.merge(
                attenuation_df[['reach_id', 'attenuation']],
                on='reach_id', how='left'
            )

        results['attenuation'].fillna(Config.DEFAULT_ATTENUATION,
                                      inplace=True)

        # For simplification, if detailed reach network not available,
        # apply routing based on HYDSEQ order
        if reach_network_df is not None:
            results = results.sort_values('HYDSEQ', ascending=True)

        # Calculate routed loads for each scenario
        for scenario in ['baseline', 'wetland', 'cw']:
            col_name = f'generated_{scenario}'
            routed_col = f'routed_{scenario}'

            if col_name in results.columns:
                # Simplified routing: assume some accumulation factor
                # In full implementation, would track upstream connectivity
                accumulation_factor = 1.0  # Start with direct load
                results[routed_col] = (results[col_name] *
                                      (1 + accumulation_factor *
                                       results['attenuation']))
            else:
                results[routed_col] = 0

        # Calculate routed reductions
        results['routed_reduction'] = (results['routed_baseline'] -
                                       results['routed_cw'])
        results['routed_reduction_percent'] = (
            (results['routed_reduction'] / results['routed_baseline']) * 100
        ).replace([np.inf, -np.inf], 0).fillna(0)

        total_baseline = results['routed_baseline'].sum()
        total_cw = results['routed_cw'].sum()
        total_reduction = total_baseline - total_cw

        print(f"  Routed baseline: {total_baseline:.4f} t/y")
        print(f"  Routed with CW:  {total_cw:.4f} t/y")
        print(f"  Total reduction: {total_reduction:.4f} t/y ({(total_reduction/total_baseline)*100:.1f}%)")

        return results

    @staticmethod
    def route_loads_advanced(loads_df, reach_network_df, attenuation_df=None):
        """
        Advanced routing with full network traversal.

        This implements the full routing algorithm:
        For each reach (in HYDSEQ order):
            Routed[i] = Generated[i] + Σ(Upstream_Routed[j] × Attenuation[i])

        Args:
            loads_df: DataFrame with generated loads
            reach_network_df: DataFrame with FROM_NODE -> TO_NODE mapping
            attenuation_df: DataFrame with reach-specific attenuation

        Returns:
            DataFrame with fully routed loads
        """
        print(f"\nPerforming advanced network routing...")

        results = loads_df.copy()

        # Build upstream connectivity
        upstream_map = defaultdict(list)
        downstream_map = {}

        if reach_network_df is not None:
            for _, row in reach_network_df.iterrows():
                from_reach = row.get('from_reach', row.get('FROM_REACH'))
                to_reach = row.get('to_reach', row.get('TO_REACH'))

                if from_reach and to_reach:
                    upstream_map[to_reach].append(from_reach)
                    downstream_map[from_reach] = to_reach

        # Load attenuation factors
        attenuation = {}
        if attenuation_df is not None:
            for _, row in attenuation_df.iterrows():
                reach_id = row.get('reach_id', row.get('NZSEGMENT'))
                atten = row.get('attenuation', row.get('PstreamCarry'))
                if reach_id and atten:
                    attenuation[reach_id] = atten

        # Route each scenario
        for scenario in ['baseline', 'wetland', 'cw']:
            gen_col = f'generated_{scenario}'
            routed_col = f'routed_{scenario}'

            if gen_col not in results.columns:
                continue

            # Initialize routed loads with generated loads
            routed = results[['reach_id', gen_col]].copy()
            routed.rename(columns={gen_col: 'value'}, inplace=True)
            routed_dict = dict(zip(routed['reach_id'], routed['value']))

            # Sort by HYDSEQ if available
            if 'HYDSEQ' in results.columns:
                sorted_reaches = results.sort_values('HYDSEQ')['reach_id'].tolist()
            else:
                sorted_reaches = results['reach_id'].tolist()

            # Route through network
            for reach_id in sorted_reaches:
                if reach_id in upstream_map:
                    # Add contributions from upstream reaches
                    upstream_contribution = 0
                    for upstream_reach in upstream_map[reach_id]:
                        if upstream_reach in routed_dict:
                            atten_factor = attenuation.get(reach_id,
                                                          Config.DEFAULT_ATTENUATION)
                            upstream_contribution += (routed_dict[upstream_reach] *
                                                    atten_factor)

                    routed_dict[reach_id] += upstream_contribution

            # Update results
            results[routed_col] = results['reach_id'].map(routed_dict)

        # Calculate reductions
        if all(col in results.columns for col in
               ['routed_baseline', 'routed_cw']):
            results['routed_reduction'] = (results['routed_baseline'] -
                                          results['routed_cw'])
            results['routed_reduction_percent'] = (
                (results['routed_reduction'] /
                 results['routed_baseline'].replace(0, np.nan)) * 100
            ).fillna(0)

        return results


# ============================================================================
# SECTION 5: RESULTS GENERATION AND OUTPUT
# ============================================================================

class ResultsGenerator:
    """Generate output files, visualizations, and summaries"""

    @staticmethod
    def create_output_directories():
        """Create output directory structure"""
        print(f"\nCreating output directories...")

        for directory in [Config.DATA_DIR, Config.FIGURES_DIR, Config.MAPS_DIR,
                         Config.SUMMARY_DIR, Config.OUTPUT_DIR]:
            Path(directory).mkdir(parents=True, exist_ok=True)
            print(f"  Created: {directory}")

    @staticmethod
    def save_results_csv(results_df, filename, output_dir=None):
        """
        Save results to CSV file with rounded numeric values.

        Args:
            results_df: DataFrame to save
            filename: Output filename
            output_dir: Output directory (uses Config.DATA_DIR if None)
        """
        if output_dir is None:
            output_dir = Config.DATA_DIR

        # Create a copy to avoid modifying original
        df_to_save = results_df.copy()

        # Round numeric columns to 4 decimal places
        numeric_cols = df_to_save.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            df_to_save[col] = df_to_save[col].round(4)

        filepath = os.path.join(output_dir, filename)
        df_to_save.to_csv(filepath, index=False)
        print(f"  Saved: {filepath}")

        return filepath

    @staticmethod
    def save_lake_omapere_excel(results_df, cw_coverage_df, filename='Lake_Omapere_CW_Analysis_DETAILED.xlsx', output_dir=None):
        """
        Create a nicely formatted Excel file for Lake Omapere reaches only.

        Args:
            results_df: Full results DataFrame
            cw_coverage_df: CW coverage DataFrame with Lake Omapere reach IDs
            filename: Output filename
            output_dir: Output directory (uses Config.DATA_DIR if None)

        Returns:
            Path to saved Excel file
        """
        if not OPENPYXL_AVAILABLE:
            print("  Warning: openpyxl not available, skipping Excel export")
            return None

        if output_dir is None:
            output_dir = Config.DATA_DIR

        print(f"\n  Creating formatted Excel file for Lake Omapere reaches...")

        # Get Lake Omapere reach IDs
        lake_reach_ids = set(cw_coverage_df['reach_id'].values)

        # Filter to Lake Omapere reaches only
        lake_results = results_df[results_df['reach_id'].isin(lake_reach_ids)].copy()

        # Sort by reach_id
        lake_results = lake_results.sort_values('reach_id')

        # Select and rename columns for clarity - base columns
        output_columns = {
            'reach_id': 'NZSEGMENT',
            'generated_baseline': 'Baseline_Load_tpy',
            'generated_wetland': 'Wetland_Load_tpy',
            'generated_cw': 'With_CW_Load_tpy',
            'cw_reduction': 'CW_Reduction_tpy',
            'CW_Coverage_Percent': 'CW_Coverage_%',
            'coverage_category': 'Coverage_Category',
            'lrf_factor': 'LRF',
            'clay_percent': 'Clay_%',
            'HighClay': 'Clay_Blocked',
            'routed_baseline': 'Routed_Baseline_tpy',
            'routed_wetland': 'Routed_Wetland_tpy',
            'routed_cw': 'Routed_CW_tpy',
            'routed_reduction': 'Routed_Reduction_tpy',
            'routed_reduction_percent': 'Routed_Reduction_%'
        }

        # Add P fraction columns if available
        p_fract_cols = {}
        for col in lake_results.columns:
            if any(fract in col for fract in ['PartP', 'DRP', 'DOP']):
                # Include main P fractions and BE columns (bank erosion split)
                # Exclude other pathway-specific columns (SR, SD, TD, IF, SG, DG) for summary
                if not any(path in col for path in ['_SR', '_SD', '_TD', '_IF', '_SG', '_DG']):
                    p_fract_cols[col] = col.replace('generated_', '').replace('_', ' ')

        # Add land use columns if available
        landuse_cols = {}
        for col in ['Pasture', 'DAIRY', 'SBINTEN', 'NAT_FOR', 'SCRUB', 'URBAN']:
            if col in lake_results.columns:
                landuse_cols[col] = f'{col}_%'

        # Add HYPE pathway columns if available
        hype_cols = {}
        for col in ['SR', 'IF', 'SG', 'DG', 'TD', 'SD']:
            if col in lake_results.columns:
                hype_cols[col] = f'{col}_Pathway_%'

        # Combine all columns
        all_columns = {**output_columns, **p_fract_cols, **landuse_cols, **hype_cols}

        # Only include columns that exist in lake_results
        available_cols = {k: v for k, v in all_columns.items() if k in lake_results.columns}

        # Create output dataframe with renamed columns
        export_df = lake_results[list(available_cols.keys())].rename(columns=available_cols)

        # Calculate additional useful columns
        export_df['Lake_Rise_Effect_tpy'] = export_df['Baseline_Load_tpy'] - export_df['Wetland_Load_tpy']
        export_df['Lake_Rise_Effect_%'] = (export_df['Lake_Rise_Effect_tpy'] / export_df['Baseline_Load_tpy'] * 100).round(2)
        export_df['CW_Reduction_%'] = (export_df['CW_Reduction_tpy'] / export_df['Wetland_Load_tpy'] * 100).round(2)
        export_df['Total_Reduction_tpy'] = export_df['Baseline_Load_tpy'] - export_df['With_CW_Load_tpy']
        export_df['Total_Reduction_%'] = (export_df['Total_Reduction_tpy'] / export_df['Baseline_Load_tpy'] * 100).round(2)
        export_df['CW_Status'] = export_df.apply(
            lambda row: 'No CW' if row['CW_Coverage_%'] == 0
                        else 'Clay Blocked' if row['Clay_Blocked']
                        else 'Effective', axis=1
        )

        # Reorder columns logically
        final_columns = [
            'NZSEGMENT',
            'CW_Coverage_%',
            'Coverage_Category',
            'LRF',
            'Clay_%',
            'Clay_Blocked',
            'CW_Status',
            'Baseline_Load_tpy',
            'Wetland_Load_tpy',
            'With_CW_Load_tpy',
            'Lake_Rise_Effect_tpy',
            'Lake_Rise_Effect_%',
            'CW_Reduction_tpy',
            'CW_Reduction_%',
            'Total_Reduction_tpy',
            'Total_Reduction_%',
            'Routed_Baseline_tpy',
            'Routed_Wetland_tpy',
            'Routed_CW_tpy',
            'Routed_Reduction_tpy',
            'Routed_Reduction_%'
        ]

        # Add P fractions, land use, and pathways columns to final list
        for col in export_df.columns:
            if col not in final_columns:
                final_columns.append(col)

        export_df = export_df[final_columns]

        # Save to Excel with formatting
        filepath = os.path.join(output_dir, filename)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lake Omapere CW Analysis"

            # Write title and metadata
            ws['A1'] = 'Lake Omapere CW Mitigation Effectiveness Analysis'
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f'Analysis Date: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A3'] = f'Total Reaches: {len(export_df)}'
            ws['A4'] = f'Reaches with CW: {(export_df["CW_Coverage_%"] > 0).sum()}'
            ws['A5'] = f'Effective CW Reaches: {(export_df["CW_Status"] == "Effective").sum()}'

            # Write data starting at row 7
            start_row = 7

            # Write headers
            for c_idx, col_name in enumerate(export_df.columns, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Write data rows
            for r_idx, row in enumerate(export_df.values, start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Apply number formatting
                    col_name = export_df.columns[c_idx - 1]
                    if '_tpy' in col_name:
                        cell.number_format = '0.0000'
                    elif '_%' in col_name or col_name == 'Clay_%' or col_name == 'CW_Coverage_%':
                        cell.number_format = '0.00'
                    elif col_name == 'LRF':
                        cell.number_format = '0.00'

                    # Conditional formatting for CW_Status
                    if col_name == 'CW_Status':
                        if value == 'Effective':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif value == 'Clay Blocked':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                        elif value == 'No CW':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="9C5700")

                    # Highlight high reductions
                    if col_name == 'CW_Reduction_tpy' and isinstance(value, (int, float)) and value > 0.05:
                        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            # Set column widths
            column_widths = {
                'A': 12,  # NZSEGMENT
                'B': 12,  # CW_Coverage_%
                'C': 15,  # Coverage_Category
                'D': 8,   # LRF
                'E': 10,  # Clay_%
                'F': 12,  # Clay_Blocked
                'G': 15,  # CW_Status
                'H': 16,  # Baseline_Load_tpy
                'I': 16,  # Wetland_Load_tpy
                'J': 16,  # With_CW_Load_tpy
                'K': 18,  # Lake_Rise_Effect_tpy
                'L': 16,  # Lake_Rise_Effect_%
                'M': 18,  # CW_Reduction_tpy
                'N': 14,  # CW_Reduction_%
                'O': 18,  # Total_Reduction_tpy
                'P': 16,  # Total_Reduction_%
                'Q': 18,  # Routed_Baseline_tpy
                'R': 18,  # Routed_Wetland_tpy
                'S': 16,  # Routed_CW_tpy
                'T': 20,  # Routed_Reduction_tpy
                'U': 18   # Routed_Reduction_%
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Freeze panes (freeze header row)
            ws.freeze_panes = ws['A8']

            # Add summary sheet
            ws_summary = wb.create_sheet("Summary Statistics")
            ws_summary['A1'] = 'Summary Statistics - Lake Omapere CW Analysis'
            ws_summary['A1'].font = Font(size=14, bold=True)

            summary_data = [
                ['Metric', 'Value', 'Unit'],
                ['', '', ''],
                ['LOADS', '', ''],
                ['Baseline TP Load', export_df['Baseline_Load_tpy'].sum(), 't/y'],
                ['Wetland TP Load', export_df['Wetland_Load_tpy'].sum(), 't/y'],
                ['With CW TP Load', export_df['With_CW_Load_tpy'].sum(), 't/y'],
                ['', '', ''],
                ['REDUCTIONS', '', ''],
                ['Lake Rise Reduction', export_df['Lake_Rise_Effect_tpy'].sum(), 't/y'],
                ['CW Reduction', export_df['CW_Reduction_tpy'].sum(), 't/y'],
                ['Total Reduction', export_df['Total_Reduction_tpy'].sum(), 't/y'],
                ['', '', ''],
                ['PERCENTAGES', '', ''],
                ['Lake Rise Effect', (export_df['Lake_Rise_Effect_tpy'].sum() / export_df['Baseline_Load_tpy'].sum() * 100), '%'],
                ['CW Mitigation', (export_df['CW_Reduction_tpy'].sum() / export_df['Wetland_Load_tpy'].sum() * 100), '%'],
                ['Total Reduction', (export_df['Total_Reduction_tpy'].sum() / export_df['Baseline_Load_tpy'].sum() * 100), '%'],
                ['', '', ''],
                ['CW IMPLEMENTATION', '', ''],
                ['Total Reaches', len(export_df), 'reaches'],
                ['Reaches with CW', (export_df['CW_Coverage_%'] > 0).sum(), 'reaches'],
                ['Effective CW Reaches', (export_df['CW_Status'] == 'Effective').sum(), 'reaches'],
                ['Clay Blocked Reaches', (export_df['CW_Status'] == 'Clay Blocked').sum(), 'reaches'],
                ['No CW Reaches', (export_df['CW_Status'] == 'No CW').sum(), 'reaches'],
            ]

            for r_idx, row in enumerate(summary_data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(size=14, bold=True)
                    elif c_idx == 1 and value in ['LOADS', 'REDUCTIONS', 'PERCENTAGES', 'CW IMPLEMENTATION']:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

                    # Number formatting
                    if c_idx == 2 and isinstance(value, (int, float)):
                        if r_idx >= 13 and r_idx <= 16:  # Percentages
                            cell.number_format = '0.00'
                        elif r_idx >= 3 and r_idx <= 11:  # Loads and reductions
                            cell.number_format = '0.0000'

            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 15
            ws_summary.column_dimensions['C'].width = 10

            # Save workbook
            wb.save(filepath)
            print(f"  Saved formatted Excel: {filepath}")
            print(f"    - Sheet 1: Lake Omapere CW Analysis ({len(export_df)} reaches)")
            print(f"    - Sheet 2: Summary Statistics")

            return filepath

        except Exception as e:
            print(f"  Error creating formatted Excel: {e}")
            # Fallback to simple Excel export
            export_df.to_excel(filepath, index=False, sheet_name='Lake Omapere CW Analysis')
            print(f"  Saved simple Excel: {filepath}")
            return filepath

    @staticmethod
    def generate_summary_statistics(results_df):
        """
        Generate summary statistics.

        Args:
            results_df: Results DataFrame

        Returns:
            Dictionary with summary statistics
        """
        summary = {
            'timestamp': datetime.now().isoformat(),
            'total_reaches': len(results_df),
            'reaches_with_cw': (results_df['CW_Coverage_Percent'] > 0).sum(),

            'generated_baseline': {
                'total': results_df['generated_baseline'].sum(),
                'mean': results_df['generated_baseline'].mean(),
                'std': results_df['generated_baseline'].std()
            },

            'generated_wetland': {
                'total': results_df['generated_wetland'].sum(),
                'mean': results_df['generated_wetland'].mean(),
                'std': results_df['generated_wetland'].std()
            },

            'generated_cw': {
                'total': results_df['generated_cw'].sum(),
                'mean': results_df['generated_cw'].mean(),
                'std': results_df['generated_cw'].std()
            },

            'generated_reduction': {
                'total': (results_df['generated_baseline'] -
                         results_df['generated_cw']).sum(),
                'percent': ((results_df['generated_baseline'].sum() -
                            results_df['generated_cw'].sum()) /
                           results_df['generated_baseline'].sum() * 100)
            },

            'cw_statistics': {
                'total_coverage_area': results_df['CW_Coverage_Percent'].sum(),
                'mean_coverage': results_df['CW_Coverage_Percent'].mean(),
                'max_coverage': results_df['CW_Coverage_Percent'].max(),
                'coverage_high': (results_df['coverage_category'] == 'high').sum(),
                'coverage_medium': (results_df['coverage_category'] == 'medium').sum(),
                'coverage_low': (results_df['coverage_category'] == 'low').sum()
            }
        }

        # Add routed statistics if available
        if 'routed_baseline' in results_df.columns:
            summary['routed_baseline'] = {
                'total': results_df['routed_baseline'].sum(),
                'mean': results_df['routed_baseline'].mean()
            }
            summary['routed_cw'] = {
                'total': results_df['routed_cw'].sum(),
                'mean': results_df['routed_cw'].mean()
            }
            summary['routed_reduction'] = {
                'total': (results_df['routed_baseline'].sum() -
                         results_df['routed_cw'].sum()),
                'percent': ((results_df['routed_baseline'].sum() -
                            results_df['routed_cw'].sum()) /
                           results_df['routed_baseline'].sum() * 100)
            }

        return summary

    @staticmethod
    def save_summary_json(summary_stats, filename='analysis_summary.json'):
        """Save summary statistics to JSON"""
        filepath = os.path.join(Config.SUMMARY_DIR, filename)

        with open(filepath, 'w') as f:
            json.dump(summary_stats, f, indent=2, default=str)

        print(f"  Saved summary: {filepath}")

    @staticmethod
    def save_summary_text(summary_stats, filename='analysis_summary.txt'):
        """Save summary statistics as formatted text"""
        filepath = os.path.join(Config.SUMMARY_DIR, filename)

        with open(filepath, 'w') as f:
            f.write("Lake Omapere CW Mitigation Analysis Summary\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"Analysis Date: {summary_stats['timestamp']}\n")
            f.write(f"Total Reaches Analyzed: {summary_stats['total_reaches']}\n")
            f.write(f"Reaches with CW: {summary_stats['reaches_with_cw']}\n\n")

            f.write("GENERATED LOADS (t/y)\n")
            f.write("-" * 60 + "\n")
            f.write(f"Baseline:\n")
            f.write(f"  Total: {summary_stats['generated_baseline']['total']:.4f}\n")
            f.write(f"  Mean:  {summary_stats['generated_baseline']['mean']:.6f}\n")

            f.write(f"\nWetland:\n")
            f.write(f"  Total: {summary_stats['generated_wetland']['total']:.4f}\n")
            f.write(f"  Mean:  {summary_stats['generated_wetland']['mean']:.6f}\n")

            f.write(f"\nWith CW Mitigation:\n")
            f.write(f"  Total: {summary_stats['generated_cw']['total']:.4f}\n")
            f.write(f"  Mean:  {summary_stats['generated_cw']['mean']:.6f}\n")

            f.write(f"\nReduction from CW:\n")
            f.write(f"  Total: {summary_stats['generated_reduction']['total']:.4f} t/y\n")
            f.write(f"  Percent: {summary_stats['generated_reduction']['percent']:.2f}%\n")

            if 'routed_baseline' in summary_stats:
                f.write("\n\nROUTED LOADS (t/y)\n")
                f.write("-" * 60 + "\n")
                f.write(f"Baseline: {summary_stats['routed_baseline']['total']:.4f}\n")
                f.write(f"With CW:  {summary_stats['routed_cw']['total']:.4f}\n")
                f.write(f"Reduction: {summary_stats['routed_reduction']['total']:.4f} t/y\n")
                f.write(f"Percent: {summary_stats['routed_reduction']['percent']:.2f}%\n")

                amplification = (summary_stats['routed_reduction']['total'] /
                               summary_stats['generated_reduction']['total'])
                f.write(f"\nRouting Amplification Factor: {amplification:.1f}×\n")

            f.write("\n\nCW COVERAGE STATISTICS\n")
            f.write("-" * 60 + "\n")
            cw_stats = summary_stats['cw_statistics']
            f.write(f"Total Coverage: {cw_stats['total_coverage_area']:.2f}%\n")
            f.write(f"Mean Coverage: {cw_stats['mean_coverage']:.2f}%\n")
            f.write(f"Max Coverage: {cw_stats['max_coverage']:.2f}%\n")
            f.write(f"High Coverage (>4%): {cw_stats['coverage_high']} reaches\n")
            f.write(f"Medium Coverage (2-4%): {cw_stats['coverage_medium']} reaches\n")
            f.write(f"Low Coverage (<2%): {cw_stats['coverage_low']} reaches\n")

        print(f"  Saved text summary: {filepath}")

    @staticmethod
    def generate_visualizations(results_df, summary_stats):
        """
        Generate visualization charts.

        Creates:
        - Load reduction comparison chart
        - CW coverage distribution
        - Generated vs Routed comparison
        """
        if not MATPLOTLIB_AVAILABLE:
            print("  Warning: matplotlib not available, skipping visualizations")
            return

        print(f"\nGenerating visualizations...")

        # Chart 1: Generated vs Routed Comparison
        if all(col in results_df.columns for col in
               ['generated_baseline', 'generated_cw',
                'routed_baseline', 'routed_cw']):

            fig, axes = plt.subplots(2, 2, figsize=(14, 10))
            fig.suptitle('Lake Omapere CW Mitigation Analysis', fontsize=16, fontweight='bold')

            # Subplot 1: Generated Baseline
            ax = axes[0, 0]
            results_df.sort_values('generated_baseline', ascending=False).head(10).plot(
                x='reach_id', y='generated_baseline', kind='bar', ax=ax, legend=False)
            ax.set_title('Top 10 Reaches: Generated Baseline Load')
            ax.set_ylabel('TP Load (t/y)')
            ax.set_xlabel('')
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')

            # Subplot 2: CW Reduction
            ax = axes[0, 1]
            reduction_data = results_df[results_df['cw_reduction'] > 0].sort_values(
                'cw_reduction', ascending=False).head(10)
            reduction_data.plot(x='reach_id', y='cw_reduction', kind='bar',
                               ax=ax, legend=False, color='green')
            ax.set_title('Top 10 Reaches: CW Reduction (Generated)')
            ax.set_ylabel('Reduction (t/y)')
            ax.set_xlabel('')
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')

            # Subplot 3: Coverage Distribution
            ax = axes[1, 0]
            coverage_counts = {
                'High (>4%)': (results_df['coverage_category'] == 'high').sum(),
                'Medium (2-4%)': (results_df['coverage_category'] == 'medium').sum(),
                'Low (<2%)': (results_df['coverage_category'] == 'low').sum(),
                'None': (results_df['coverage_category'] == 'none').sum()
            }
            ax.bar(coverage_counts.keys(), coverage_counts.values(), color=['#2ecc71', '#f39c12', '#e74c3c', '#95a5a6'])
            ax.set_title('CW Coverage Distribution (Reaches)')
            ax.set_ylabel('Number of Reaches')
            ax.set_xlabel('Coverage Category')
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')

            # Subplot 4: Total Reduction Summary
            ax = axes[1, 1]
            scenarios = ['Baseline', 'With CW']
            generated = [results_df['generated_baseline'].sum(),
                        results_df['generated_cw'].sum()]
            routed = [results_df['routed_baseline'].sum(),
                     results_df['routed_cw'].sum()]

            x = np.arange(len(scenarios))
            width = 0.35

            ax.bar(x - width/2, generated, width, label='Generated', color='#3498db')
            ax.bar(x + width/2, routed, width, label='Routed', color='#e67e22')

            ax.set_ylabel('TP Load (t/y)')
            ax.set_title('Total Load Comparison')
            ax.set_xticks(x)
            ax.set_xticklabels(scenarios)
            ax.legend()

            plt.tight_layout()
            filepath = os.path.join(Config.FIGURES_DIR,
                                   'CW_Analysis_Summary.png')
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            print(f"  Saved: {filepath}")
            plt.close()

        # Chart 2: Reduction Percentages
        if 'routed_reduction_percent' in results_df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))

            top_reductions = results_df.nlargest(15, 'routed_reduction_percent')
            ax.barh(range(len(top_reductions)), top_reductions['routed_reduction_percent'],
                   color='#27ae60')
            ax.set_yticks(range(len(top_reductions)))
            ax.set_yticklabels(top_reductions['reach_id'].astype(str))
            ax.set_xlabel('Reduction (%)')
            ax.set_title('Top 15 Reaches by Routed Load Reduction (%)')
            ax.invert_yaxis()

            plt.tight_layout()
            filepath = os.path.join(Config.FIGURES_DIR,
                                   'Reduction_Percent_Top_Reaches.png')
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            print(f"  Saved: {filepath}")
            plt.close()


# ============================================================================
# SECTION 5B: SPATIAL MAPPING
# ============================================================================

class MapGenerator:
    """Generate spatial maps of phosphorus loads across river network"""

    @staticmethod
    def load_river_network(shapefile_path):
        """
        Load river network shapefile.

        Args:
            shapefile_path: Path to river shapefile

        Returns:
            GeoDataFrame with river geometry
        """
        if not GEOPANDAS_AVAILABLE:
            print("  Warning: geopandas not available, skipping mapping")
            return None

        print(f"\nLoading river network from: {shapefile_path}")

        if not os.path.exists(shapefile_path):
            print(f"  Warning: Shapefile not found: {shapefile_path}")
            return None

        try:
            gdf = gpd.read_file(shapefile_path)
            print(f"  Loaded {len(gdf)} river reaches")
            return gdf
        except Exception as e:
            print(f"  Error loading shapefile: {e}")
            return None

    @staticmethod
    def filter_lake_reaches(river_gdf, results_df, reach_id_col='NZSEGMENT'):
        """
        Filter river network to Lake Omapere reaches only.

        Args:
            river_gdf: GeoDataFrame with all river reaches
            results_df: DataFrame with analysis results
            reach_id_col: Column name for reach ID

        Returns:
            GeoDataFrame filtered to Lake reaches with results joined
        """
        if river_gdf is None:
            return None

        print(f"\nFiltering to Lake Omapere reaches...")

        # Get list of Lake reach IDs from results
        lake_reach_ids = set(results_df['reach_id'].values)

        # Filter river geometry to Lake reaches
        if reach_id_col not in river_gdf.columns:
            # Try common alternatives
            possible_cols = [col for col in river_gdf.columns
                           if 'segment' in col.lower() or 'nzseg' in col.lower()]
            if possible_cols:
                reach_id_col = possible_cols[0]
                print(f"  Using column: {reach_id_col}")
            else:
                print(f"  Error: Cannot find reach ID column in shapefile")
                return None

        # Filter to Lake reaches
        lake_gdf = river_gdf[river_gdf[reach_id_col].isin(lake_reach_ids)].copy()

        # Join with results data
        lake_gdf = lake_gdf.merge(
            results_df,
            left_on=reach_id_col,
            right_on='reach_id',
            how='left'
        )

        print(f"  Filtered to {len(lake_gdf)} Lake reaches")
        return lake_gdf

    @staticmethod
    def create_phosphorus_map(lake_gdf, value_column, title, output_path,
                             cmap='RdYlGn_r', vmin=None, vmax=None,
                             catchment_gdf=None, lake_gdf_poly=None, legend_label=None):
        """
        Create map of phosphorus loads across river network.

        Args:
            lake_gdf: GeoDataFrame with Lake reaches and values
            value_column: Column name to map
            title: Map title
            output_path: Path to save PNG
            cmap: Colormap name
            vmin: Minimum value for color scale
            vmax: Maximum value for color scale
            catchment_gdf: Optional catchment boundary
            lake_gdf_poly: Optional lake polygon
            legend_label: Optional legend label (auto-detected if None)

        Returns:
            Path to saved map
        """
        if not MATPLOTLIB_AVAILABLE or not GEOPANDAS_AVAILABLE:
            print("  Warning: matplotlib or geopandas not available")
            return None

        if lake_gdf is None or len(lake_gdf) == 0:
            print(f"  Warning: No data to map")
            return None

        print(f"\n  Creating map: {title}")

        try:
            # Create figure
            fig, ax = plt.subplots(figsize=(10, 12))

            # Plot catchment boundary if available
            if catchment_gdf is not None:
                catchment_gdf.boundary.plot(ax=ax, color='gray', linewidth=1.5,
                                          label='Catchment')

            # Plot lake if available
            if lake_gdf_poly is not None:
                lake_gdf_poly.plot(ax=ax, color='lightblue', alpha=0.5,
                                  edgecolor='blue', linewidth=1, label='Lake')

            # Get value range
            if value_column not in lake_gdf.columns:
                print(f"  Error: Column '{value_column}' not found")
                return None

            values = lake_gdf[value_column].dropna()
            if len(values) == 0:
                print(f"  Warning: No valid values in '{value_column}'")
                return None

            if vmin is None:
                vmin = values.min()
            if vmax is None:
                vmax = values.max()

            print(f"    Value range: {vmin:.4f} to {vmax:.4f}")

            # Auto-detect legend label if not provided
            if legend_label is None:
                if 'coverage' in value_column.lower() or 'percent' in value_column.lower():
                    legend_label = 'Coverage (%)'
                elif 'reduction' in value_column.lower():
                    legend_label = 'TP Reduction (t/y)'
                else:
                    legend_label = 'TP Load (t/y)'

            # Plot river reaches with color by value
            lake_gdf.plot(column=value_column,
                         ax=ax,
                         cmap=cmap,
                         linewidth=2.5,
                         vmin=vmin,
                         vmax=vmax,
                         legend=True,
                         legend_kwds={
                             'label': legend_label,
                             'orientation': 'horizontal',
                             'shrink': 0.8,
                             'pad': 0.05
                         })

            # Format axes
            ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
            ax.set_xlabel('Longitude', fontsize=10)
            ax.set_ylabel('Latitude', fontsize=10)
            ax.tick_params(labelsize=8)

            # Add grid
            ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)

            # Tight layout
            plt.tight_layout()

            # Save
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"    Saved: {output_path}")
            plt.close()

            return output_path

        except Exception as e:
            print(f"  Error creating map: {e}")
            plt.close()
            return None

    @staticmethod
    def create_comparison_map(lake_gdf, scenarios, output_path,
                             catchment_gdf=None, lake_gdf_poly=None):
        """
        Create multi-panel comparison map for different scenarios.

        Args:
            lake_gdf: GeoDataFrame with Lake reaches
            scenarios: List of (column_name, title) tuples
            output_path: Path to save PNG
            catchment_gdf: Optional catchment boundary
            lake_gdf_poly: Optional lake polygon

        Returns:
            Path to saved map
        """
        if not MATPLOTLIB_AVAILABLE or not GEOPANDAS_AVAILABLE:
            return None

        if lake_gdf is None or len(scenarios) == 0:
            return None

        print(f"\n  Creating comparison map with {len(scenarios)} scenarios")

        try:
            # Create figure with subplots
            n_scenarios = len(scenarios)
            n_cols = 2 if n_scenarios > 1 else 1
            n_rows = (n_scenarios + 1) // 2

            fig, axes = plt.subplots(n_rows, n_cols, figsize=(16, 6 * n_rows))

            if n_scenarios == 1:
                axes = [axes]
            else:
                axes = axes.flatten()

            # Get global min/max for consistent coloring
            all_values = []
            for col, _ in scenarios:
                if col in lake_gdf.columns:
                    all_values.extend(lake_gdf[col].dropna().values)

            if len(all_values) == 0:
                print("  Warning: No valid values to map")
                return None

            vmin = min(all_values)
            vmax = max(all_values)

            print(f"    Global value range: {vmin:.4f} to {vmax:.4f}")

            # Create each subplot
            for idx, (col, title) in enumerate(scenarios):
                ax = axes[idx]

                # Plot catchment
                if catchment_gdf is not None:
                    catchment_gdf.boundary.plot(ax=ax, color='gray', linewidth=1)

                # Plot lake
                if lake_gdf_poly is not None:
                    lake_gdf_poly.plot(ax=ax, color='lightblue', alpha=0.3,
                                      edgecolor='blue', linewidth=0.8)

                # Plot river reaches
                if col in lake_gdf.columns:
                    lake_gdf.plot(column=col,
                                 ax=ax,
                                 cmap='RdYlGn_r',
                                 linewidth=2.5,
                                 vmin=vmin,
                                 vmax=vmax,
                                 legend=False)

                ax.set_title(title, fontsize=12, fontweight='bold')
                ax.set_xlabel('')
                ax.set_ylabel('')
                ax.tick_params(labelsize=8)
                ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)

            # Hide extra subplots
            for idx in range(n_scenarios, len(axes)):
                axes[idx].set_visible(False)

            # Add colorbar
            sm = plt.cm.ScalarMappable(
                cmap='RdYlGn_r',
                norm=plt.Normalize(vmin=vmin, vmax=vmax)
            )
            sm.set_array([])

            # Place colorbar at bottom
            cbar_ax = fig.add_axes([0.15, 0.05, 0.7, 0.02])
            cbar = fig.colorbar(sm, cax=cbar_ax, orientation='horizontal')
            cbar.set_label('TP Load (t/y)', fontsize=10)

            # Overall title
            fig.suptitle('Lake Omapere TP Loads - Scenario Comparison',
                        fontsize=16, fontweight='bold', y=0.98)

            plt.tight_layout(rect=[0, 0.08, 1, 0.96])

            # Save
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"    Saved: {output_path}")
            plt.close()

            return output_path

        except Exception as e:
            print(f"  Error creating comparison map: {e}")
            plt.close()
            return None

    @staticmethod
    def generate_all_maps(results_df):
        """
        Generate all phosphorus maps.

        Creates:
        - Baseline generated load map
        - Wetland generated load map
        - CW mitigation generated load map
        - Baseline routed load map
        - CW mitigation routed load map
        - Multi-panel comparison maps

        Args:
            results_df: DataFrame with analysis results

        Returns:
            List of created map paths
        """
        print("\n[GENERATING SPATIAL MAPS]")
        print("-" * 70)

        if not GEOPANDAS_AVAILABLE:
            print("  Skipping mapping - geopandas not available")
            print("  Install with: pip install geopandas")
            return []

        created_maps = []

        # Create maps directory
        Path(Config.MAPS_DIR).mkdir(parents=True, exist_ok=True)

        try:
            # Load river network
            river_gdf = MapGenerator.load_river_network(Config.RIVER_SHAPEFILE)
            if river_gdf is None:
                return []

            # Filter to Lake reaches and join with results
            lake_gdf = MapGenerator.filter_lake_reaches(river_gdf, results_df)
            if lake_gdf is None:
                return []

            # Load optional context layers
            catchment_gdf = None
            if os.path.exists(Config.CATCHMENT_SHAPEFILE):
                try:
                    catchment_gdf = gpd.read_file(Config.CATCHMENT_SHAPEFILE)
                    print(f"  Loaded catchment boundary")
                except:
                    pass

            lake_poly_gdf = None
            if os.path.exists(Config.LAKE_SHAPEFILE):
                try:
                    lake_poly_gdf = gpd.read_file(Config.LAKE_SHAPEFILE)
                    print(f"  Loaded lake polygon")
                except:
                    pass

            # Map 1: Generated Loads Comparison
            if all(col in lake_gdf.columns for col in
                   ['generated_baseline', 'generated_wetland', 'generated_cw']):

                scenarios = [
                    ('generated_baseline', 'Baseline (Current Lake)'),
                    ('generated_wetland', 'Wetland (+0.66m, No CW)'),
                    ('generated_cw', 'Wetland + CW Mitigation')
                ]

                path = os.path.join(Config.MAPS_DIR,
                                   'Generated_Loads_Comparison.png')
                result = MapGenerator.create_comparison_map(
                    lake_gdf, scenarios, path, catchment_gdf, lake_poly_gdf)
                if result:
                    created_maps.append(result)

            # Map 2: Routed Loads Comparison
            if all(col in lake_gdf.columns for col in
                   ['routed_baseline', 'routed_cw']):

                scenarios = [
                    ('routed_baseline', 'Routed Baseline'),
                    ('routed_cw', 'Routed with CW Mitigation')
                ]

                path = os.path.join(Config.MAPS_DIR,
                                   'Routed_Loads_Comparison.png')
                result = MapGenerator.create_comparison_map(
                    lake_gdf, scenarios, path, catchment_gdf, lake_poly_gdf)
                if result:
                    created_maps.append(result)

            # Map 3: CW Reduction (Generated)
            if 'cw_reduction' in lake_gdf.columns:
                path = os.path.join(Config.MAPS_DIR,
                                   'CW_Reduction_Generated.png')
                result = MapGenerator.create_phosphorus_map(
                    lake_gdf, 'cw_reduction',
                    'CW Mitigation Effect - Generated Loads',
                    path, cmap='Greens', vmin=0,
                    catchment_gdf=catchment_gdf,
                    lake_gdf_poly=lake_poly_gdf)
                if result:
                    created_maps.append(result)

            # Map 4: CW Reduction (Routed)
            if 'routed_reduction' in lake_gdf.columns:
                path = os.path.join(Config.MAPS_DIR,
                                   'CW_Reduction_Routed.png')
                result = MapGenerator.create_phosphorus_map(
                    lake_gdf, 'routed_reduction',
                    'CW Mitigation Effect - Routed Loads (Network)',
                    path, cmap='Greens', vmin=0,
                    catchment_gdf=catchment_gdf,
                    lake_gdf_poly=lake_poly_gdf)
                if result:
                    created_maps.append(result)

            # Map 5: Coverage Distribution
            if 'CW_Coverage_Percent' in lake_gdf.columns:
                path = os.path.join(Config.MAPS_DIR,
                                   'CW_Coverage_Distribution.png')
                result = MapGenerator.create_phosphorus_map(
                    lake_gdf, 'CW_Coverage_Percent',
                    'CW Site Coverage by Reach (%)',
                    path, cmap='YlGnBu', vmin=0,
                    catchment_gdf=catchment_gdf,
                    lake_gdf_poly=lake_poly_gdf)
                if result:
                    created_maps.append(result)

            print(f"\nOK Created {len(created_maps)} maps")
            return created_maps

        except Exception as e:
            print(f"\nERROR Error generating maps: {e}")
            return created_maps


# ============================================================================
# SECTION 6: MAIN ANALYSIS ORCHESTRATOR
# ============================================================================

class LakeOmapereAnalysis:
    """Main orchestrator for Lake Omapere CW analysis"""

    def __init__(self):
        """Initialize analysis"""
        self.baseline_clues = None
        self.wetland_clues = None
        self.cw_coverage = None
        self.clay_data = None
        self.reach_network = None
        self.attenuation = None
        self.results = None

        # NEW: Additional data for P fractions and pathways
        self.p_fractions = None
        self.landuse_data = None
        self.hype_pathways = None
        self.pathway_lrfs = None

        print("\n" + "=" * 70)
        print("Lake Omapere CW Mitigation Effectiveness Analysis")
        print("Baseline: With +0.66m lake rise effect")
        print("=" * 70)

    def load_all_data(self):
        """Load all required input data"""
        print("\n[STEP 1] LOADING INPUT DATA")
        print("-" * 70)

        loader = DataLoader()

        try:
            # Load CLUES baseline (already includes +0.66m lake rise effect)
            print("\n  [NOTE] Baseline includes +0.66m lake rise effect")
            self.baseline_clues = loader.load_clues_excel(
                Config.CLUES_BASELINE_PATH, "baseline (with +0.66m lake rise)")
            self.wetland_clues = None  # Not used - only baseline with lake rise

            # Load supporting data
            self.cw_coverage = loader.load_cw_coverage(Config.CW_COVERAGE_XLSX)
            self.clay_data = loader.load_clay_data(Config.FSL_DATA_CSV)
            self.reach_network = loader.load_reach_network(Config.REACH_NETWORK_CSV)
            self.attenuation = loader.load_attenuation_factors(Config.ATTENUATION_CSV)

            # NEW: Load P fractions, land use, and pathways data
            self.p_fractions = loader.load_p_fractions(Config.CONTAMINANT_SPLITS_XLSX)
            self.landuse_data = loader.load_landuse(Config.LANDUSE_CSV)
            self.hype_pathways = loader.load_hype_pathways(Config.HYPE_CSV)
            self.pathway_lrfs = loader.load_pathway_lrfs(Config.LRF_XLSX)

            # FILTER ALL DATA TO LAKE OMAPERE REACHES ONLY
            lake_reach_ids = self.cw_coverage['reach_id'].astype(int).tolist()
            print(f"\n[FILTER] Filtering all datasets to {len(lake_reach_ids)} Lake Omapere reaches...")

            # Filter network data
            initial_network = len(self.reach_network)
            self.reach_network = self.reach_network[self.reach_network['NZSEGMENT'].isin(lake_reach_ids)].copy()
            print(f"  Network: {initial_network} -> {len(self.reach_network)} reaches")

            # Filter attenuation data
            initial_atten = len(self.attenuation)
            self.attenuation = self.attenuation[self.attenuation['NZSEGMENT'].isin(lake_reach_ids)].copy()
            print(f"  Attenuation: {initial_atten} -> {len(self.attenuation)} reaches")

            # Filter clay data
            initial_clay = len(self.clay_data)
            self.clay_data = self.clay_data[self.clay_data['NZSEGMENT'].isin(lake_reach_ids)].copy()
            print(f"  Clay data: {initial_clay} -> {len(self.clay_data)} reaches")

            # Filter land use data
            if self.landuse_data is not None:
                initial_landuse = len(self.landuse_data)
                self.landuse_data = self.landuse_data[self.landuse_data['NZSEGMENT'].isin(lake_reach_ids)].copy()
                print(f"  Land use: {initial_landuse} -> {len(self.landuse_data)} reaches")

            # Filter HYPE pathways data
            if self.hype_pathways is not None:
                initial_hype = len(self.hype_pathways)
                self.hype_pathways = self.hype_pathways[self.hype_pathways['NZSEGMENT'].isin(lake_reach_ids)].copy()
                print(f"  HYPE pathways: {initial_hype} -> {len(self.hype_pathways)} reaches")

            print("\nOK All data loaded and filtered to Lake Omapere reaches")

        except Exception as e:
            print(f"\nERROR Error loading data: {e}")
            print("  Continuing with available data...")

    def calculate_generated_loads(self):
        """Calculate generated loads for all scenarios"""
        print("\n[STEP 2] CALCULATING GENERATED LOADS")
        print("-" * 70)

        try:
            # Extract load components (baseline includes +0.66m lake rise effect)
            calc = GeneratedLoadsCalculator()
            baseline_components = calc.extract_load_components(self.baseline_clues)

            # Calculate generated loads (single baseline scenario with +0.66m lake rise)
            self.results = baseline_components[['reach_id', 'total_load']].copy()
            self.results.rename(columns={'total_load': 'generated_baseline'}, inplace=True)

            # Also create generated_wetland column for compatibility with rest of pipeline
            self.results['generated_wetland'] = self.results['generated_baseline']

            print(f"\n  Baseline (with +0.66m lake rise) mean load: {self.results['generated_baseline'].mean():.4f} t/y")

            # FILTER TO LAKE OMAPERE REACHES ONLY
            lake_reach_ids = self.cw_coverage['reach_id'].astype(int).tolist()
            initial_count = len(self.results)
            self.results = self.results[self.results['reach_id'].isin(lake_reach_ids)].copy()
            final_count = len(self.results)

            print(f"\n  [FILTER] Filtered from {initial_count} reaches to {final_count} Lake Omapere reaches")
            print(f"  Lake Omapere total load: {self.results['generated_baseline'].sum():.4f} t/y")

            print("\nOK Generated loads calculated (baseline with +0.66m lake rise)")

        except Exception as e:
            print(f"\nERROR Error calculating generated loads: {e}")
            raise

    def apply_cw_mitigation(self):
        """Apply CW mitigation to results"""
        print("\n[STEP 3] APPLYING CW MITIGATION")
        print("-" * 70)

        try:
            cw_calc = CWMitigationCalculator()
            self.results = cw_calc.apply_cw_mitigation(
                self.results, self.cw_coverage, self.clay_data)

            print("\nOK CW mitigation applied")

        except Exception as e:
            print(f"\nERROR Error applying CW mitigation: {e}")
            raise

    def calculate_p_fractions_and_pathways(self):
        """Calculate P fractions and pathway distributions"""
        print("\n[STEP 3B] CALCULATING P FRACTIONS AND PATHWAYS")
        print("-" * 70)

        try:
            p_calc = PFractionPathwayCalculator()

            # Split TP into P fractions (PartP, DRP, DOP)
            if self.p_fractions is not None:
                self.results = p_calc.split_tp_into_fractions(
                    self.results, self.p_fractions)
                print("  P fractions calculated (PartP, DRP, DOP)")
            else:
                print("  Warning: P fractions not available, skipping")

            # PHASE 1: Calculate Bank Erosion (BE) split
            # 50% of each P fraction goes to BE pathway, 50% distributed through HYPE pathways
            self.results = p_calc.calculate_bank_erosion_split(
                self.results, sediment_pct=0.5)
            print("  Bank erosion split calculated (50% BE, 50% HYPE pathways)")

            # Split P fractions by HYPE pathways
            if self.hype_pathways is not None:
                self.results = p_calc.split_by_hype_pathways(
                    self.results, self.hype_pathways, self.p_fractions)
                print("  Pathway distributions calculated (SR, IF, SG, DG, TD, SD)")
            else:
                print("  Warning: HYPE pathways not available, skipping")

            # PHASE 2: Apply pathway-specific CW mitigation
            if self.pathway_lrfs is not None and self.p_fractions is not None:
                print("\n  Applying Phase 2 pathway-specific CW mitigation...")
                self.results = CWMitigationCalculator.apply_pathway_specific_mitigation(
                    self.results, self.pathway_lrfs, self.p_fractions)
            else:
                print("  Warning: Pathway LRFs or P fractions not available, skipping Phase 2")

            # Merge land use data
            if self.landuse_data is not None:
                self.results = self.results.merge(
                    self.landuse_data, on='reach_id', how='left')
                print("  Land use data merged")
            else:
                print("  Warning: Land use data not available, skipping")

            print("\nOK P fractions and pathways calculated")

        except Exception as e:
            print(f"\nERROR Error calculating P fractions/pathways: {e}")
            print("  Continuing without P fraction detail...")
            import traceback
            traceback.print_exc()

    def route_loads(self):
        """Route loads through network"""
        print("\n[STEP 4] ROUTING LOADS THROUGH NETWORK")
        print("-" * 70)

        try:
            router = NetworkRouter()

            # Try advanced routing if network available
            if self.reach_network is not None:
                self.results = router.route_loads_advanced(
                    self.results, self.reach_network, self.attenuation)
            else:
                self.results = router.route_loads(
                    self.results, self.reach_network, self.attenuation)

            print("\nOK Load routing completed")

        except Exception as e:
            print(f"\nERROR Error routing loads: {e}")
            print("  Continuing without routing...")
            self.results['routed_baseline'] = self.results['generated_baseline']
            self.results['routed_wetland'] = self.results['generated_wetland']
            self.results['routed_cw'] = self.results['generated_cw']

    def generate_outputs(self):
        """Generate all outputs"""
        print("\n[STEP 5] GENERATING OUTPUTS")
        print("-" * 70)

        try:
            gen = ResultsGenerator()

            # Create directories
            gen.create_output_directories()

            # Save results CSV
            gen.save_results_csv(self.results,
                                'Lake_Omapere_Analysis_Results.csv')

            # Save formatted Excel file for Lake Omapere reaches
            gen.save_lake_omapere_excel(self.results, self.cw_coverage)

            # Generate summary statistics
            summary = gen.generate_summary_statistics(self.results)
            gen.save_summary_json(summary)
            gen.save_summary_text(summary)

            # Generate visualizations
            gen.generate_visualizations(self.results, summary)

            # Generate spatial maps
            MapGenerator.generate_all_maps(self.results)

            print("\nOK All outputs generated")

            return summary

        except Exception as e:
            print(f"\nERROR Error generating outputs: {e}")
            raise

    def run_full_analysis(self):
        """Run complete analysis pipeline"""
        try:
            self.load_all_data()
            self.calculate_generated_loads()
            self.apply_cw_mitigation()
            self.calculate_p_fractions_and_pathways()  # NEW: Add P fractions and pathways
            self.route_loads()
            summary = self.generate_outputs()

            print("\n" + "=" * 70)
            print("ANALYSIS COMPLETE")
            print("=" * 70)
            print(f"\nBaseline: With +0.66m lake rise effect")
            print(f"CW Scenario: Baseline + CW mitigation")
            print(f"\nResults saved to: {Config.OUTPUT_DIR}")
            print(f"\nKey findings:")
            print(f"  CW Reduction: {summary['generated_reduction']['total']:.4f} t/y ({summary['generated_reduction']['percent']:.2f}%)")

            if 'routed_reduction' in summary:
                print(f"  CW Reduction (routed): {summary['routed_reduction']['total']:.4f} t/y ({summary['routed_reduction']['percent']:.2f}%)")
                amplification = (summary['routed_reduction']['total'] /
                               summary['generated_reduction']['total'])
                print(f"  Routing Amplification: {amplification:.1f}×")

            print("\n" + "=" * 70)

            return self.results, summary

        except Exception as e:
            print("\n" + "=" * 70)
            print("ANALYSIS FAILED")
            print("=" * 70)
            print(f"Error: {e}")
            raise


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point"""
    try:
        analysis = LakeOmapereAnalysis()
        results, summary = analysis.run_full_analysis()
        return results, summary

    except KeyboardInterrupt:
        print("\n\nAnalysis interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n\nFatal error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    results, summary = main()
