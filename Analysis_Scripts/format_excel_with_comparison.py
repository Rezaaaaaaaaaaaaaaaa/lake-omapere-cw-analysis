#!/usr/bin/env python3
"""
Format Excel Results - WITH INUNDATION COMPARISON
Applies formatting to Phase 2 results with inundation comparison columns
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

print("="*80)
print("FORMATTING EXCEL RESULTS - WITH INUNDATION COMPARISON")
print("="*80)

# Read the existing results
input_file = "Results/PHASE2_RESULTS/Lake_Omapere_CW_Analysis_PHASE2_with_comparison.xlsx"
output_file = "Results/PHASE2_RESULTS/Lake_Omapere_CW_Analysis_PHASE2_with_comparison.xlsx"

print(f"\nLoading: {input_file}")
df_results = pd.read_excel(input_file, sheet_name='Results')
df_descriptions = pd.read_excel(input_file, sheet_name='Column_Descriptions')

print(f"  Results: {len(df_results)} rows x {len(df_results.columns)} columns")

# ============================================================================
# CREATE INPUT DATA SOURCE MAPPING
# ============================================================================

print("\n" + "="*80)
print("CREATING INPUT DATA SOURCE MAPPING")
print("="*80)

# Map each column to its input data source
input_data_sources = {
    'reach_id': 'CLUES baseline file (NZSEGMENT)',
    'HYDSEQ': 'CLUES baseline file (hydrological sequence number)',
    'Scenario': 'Analysis parameter',
    'ag_percent': 'ag_percentage_by_reach.csv (calculated from LakeAreaUpdate_ASD.xlsx)',
    'ag_filter_applied': 'Calculated (TRUE if ag_percent < 25%)',
    'Total_CLUES_TP': 'CLUESloads_baseline.csv (TPAgGen + soilP + TPGen) WITH +0.66m inundation',
    'Total_TP_NoInundation': 'CLUESloads.csv (TPAgGen + soilP + TPGen) WITHOUT inundation',
    'Inundation_Reduction_TP': 'Calculated (Total_TP_NoInundation - Total_CLUES_TP)',
    'Inundation_Reduction_Percent': 'Calculated ((Inundation_Reduction_TP / Total_TP_NoInundation) x 100)',
    'Available_Load': 'Calculated (Total_CLUES_TP x ag_filter)',
    'CW_Coverage_Percent': 'CW_Coverage_GIS_CALCULATED.xlsx (Combined_Percent or Type2_SW_Percent)',
    'coverage_category': 'Calculated from CW_Coverage_Percent',
    'ExtCode': 'Calculated from coverage_category (1=SMALL, 2=MEDIUM, 3=LARGE)',
    'clay_percent': 'FSLData.csv (ClayPC column)',
    'generated_baseline': 'Calculated (Available_Load without CW mitigation)',
    'generated_with_cw': 'Calculated (Available_Load with CW mitigation applied)',
    'cw_reduction': 'Calculated (generated_baseline - generated_with_cw)',
    'cw_reduction_percent': 'Calculated (cw_reduction / generated_baseline x 100)',
    'PstreamCarry': 'AttenCarry.csv',
    'routed_baseline': 'Calculated (generated_baseline x PstreamCarry)',
    'routed_with_cw': 'Calculated (generated_with_cw x PstreamCarry)',
    'routed_reduction': 'Calculated (routed_baseline - routed_with_cw)',
    'routed_reduction_percent': 'Calculated (routed_reduction / routed_baseline x 100)',
}

# Add P fraction sources
for p_type in ['PartP', 'DRP', 'DOP']:
    input_data_sources[f'{p_type}_baseline'] = f'Calculated (Available_Load x {p_type}% split)'
    input_data_sources[f'{p_type}_bank_erosion'] = f'Calculated ({p_type}_baseline x 50%)'
    input_data_sources[f'{p_type}_hillslope'] = f'Calculated ({p_type}_baseline x 50%)'
    input_data_sources[f'{p_type}_with_cw'] = f'Calculated ({p_type}_bank_erosion + pathway remainings)'
    input_data_sources[f'{p_type}_removed'] = f'Calculated ({p_type}_baseline - {p_type}_with_cw)'

# Add pathway sources
pathways = ['SR', 'TD', 'IF', 'SG', 'DG']
for p_type in ['PartP', 'DRP', 'DOP']:
    for pathway in pathways:
        if p_type == 'PartP':
            if pathway == 'SR':
                input_data_sources[f'{p_type}_{pathway}_input'] = f'Hype.csv (100% to SR for PartP - NEW!)'
            else:
                input_data_sources[f'{p_type}_{pathway}_input'] = f'Fixed at 0% for PartP (NEW!)'
        else:
            input_data_sources[f'{p_type}_{pathway}_input'] = f'Hype.csv ({pathway} pathway %)'

        input_data_sources[f'{p_type}_{pathway}_removed'] = f'Calculated (input x LRF removal %)'
        input_data_sources[f'{p_type}_{pathway}_remaining'] = f'LRFs_years.xlsx (CW sheet, ExtCode + Pathway)'

# Add source column to descriptions
df_descriptions['Input_Data_Source'] = df_descriptions['Column'].map(input_data_sources)

# Fill any missing sources
df_descriptions['Input_Data_Source'] = df_descriptions['Input_Data_Source'].fillna('Calculated from other columns')

print(f"  Added input data sources for {len(df_descriptions)} columns")

# ============================================================================
# ROUND NUMERIC COLUMNS
# ============================================================================

print("\n" + "="*80)
print("ROUNDING NUMERIC VALUES")
print("="*80)

small_value_columns = [col for col in df_results.columns if any(x in col.lower() for x in
    ['_input', '_removed', '_remaining', '_baseline', '_with_cw', '_reduction',
     'total_clues', 'total_tp', 'inundation_reduction_tp', 'available_load', 'generated', 'routed', 'bank_erosion', 'hillslope'])]

percentage_columns = [col for col in df_results.columns if any(x in col.lower() for x in
    ['percent', 'pstreamcarry', 'clay_percent', 'cw_coverage'])]

# Round numeric columns
rounded_count = 0
for col in df_results.columns:
    if df_results[col].dtype in ['float64', 'float32']:
        if col in small_value_columns and 'percent' not in col.lower():
            # Small values (loads) - round to 5 decimals for precision
            df_results[col] = df_results[col].round(5)
            rounded_count += 1
        elif col in percentage_columns or 'percent' in col.lower():
            # Percentages - round to 2 decimals
            df_results[col] = df_results[col].round(2)
            rounded_count += 1
        else:
            # Other numeric values - round to 4 decimals
            df_results[col] = df_results[col].round(4)
            rounded_count += 1

print(f"  Rounded {rounded_count} numeric columns")

# ============================================================================
# WRITE TO EXCEL
# ============================================================================

print("\n" + "="*80)
print("WRITING TO EXCEL WITH IMPROVED FORMAT")
print("="*80)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_results.to_excel(writer, sheet_name='Results', index=False)
    df_descriptions.to_excel(writer, sheet_name='Column_Descriptions', index=False)

print(f"  Written to: {output_file}")

# ============================================================================
# APPLY ADVANCED FORMATTING
# ============================================================================

print("\n" + "="*80)
print("APPLYING ADVANCED FORMATTING")
print("="*80)

wb = load_workbook(output_file)

# Define styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Scenario colors
scenario1_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')  # Light blue
scenario2_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray

cell_alignment = Alignment(horizontal='left', vertical='center')
number_alignment = Alignment(horizontal='right', vertical='center')

border_side = Side(style='thin', color='D3D3D3')
border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

scenario_border_bottom = Side(style='medium', color='366092')
scenario_border = Border(left=border_side, right=border_side, top=border_side, bottom=scenario_border_bottom)

# Format Results sheet
print("\n  Formatting 'Results' sheet...")
ws_results = wb['Results']

# Apply header formatting
for cell in ws_results[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Freeze panes
ws_results.freeze_panes = 'C2'  # Freeze reach_id and Scenario columns

# Set column widths
print("    Setting column widths...")
for idx, col in enumerate(df_results.columns, 1):
    col_letter = get_column_letter(idx)

    # Determine width based on column content
    if col == 'reach_id':
        width = 12
    elif col == 'Scenario':
        width = 28  # Wider for scenario names
    elif 'inundation' in col.lower():
        width = 20  # Wider for inundation columns
    elif 'percent' in col.lower() or 'extcode' in col.lower():
        width = 14
    elif any(x in col.lower() for x in ['_input', '_removed', '_remaining', 'baseline', 'with_cw', 'reduction']):
        width = 16
    elif col == 'coverage_category':
        width = 18
    else:
        width = 15

    ws_results.column_dimensions[col_letter].width = width

# Apply data formatting with scenario-based coloring
print("    Applying scenario-based row coloring...")
current_scenario = None
scenario_fill = None

for row_idx in range(2, len(df_results) + 2):
    # Get scenario for this row
    scenario = df_results.iloc[row_idx - 2]['Scenario']

    # Determine fill color based on scenario
    if 'Scenario1' in str(scenario):
        scenario_fill = scenario1_fill
    else:
        scenario_fill = scenario2_fill

    # Check if this is the last row of a scenario group
    is_last_in_group = False
    if row_idx < len(df_results) + 1:
        next_scenario = df_results.iloc[row_idx - 1]['Scenario'] if row_idx - 1 < len(df_results) else None
        if next_scenario and scenario != next_scenario:
            is_last_in_group = True
    else:
        is_last_in_group = True

    # Apply formatting to all cells in row
    for col_idx, col in enumerate(df_results.columns, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws_results[f'{col_letter}{row_idx}']

        # Apply background color
        cell.fill = scenario_fill

        # Apply border (thicker at scenario boundaries)
        if is_last_in_group:
            cell.border = scenario_border
        else:
            cell.border = border

        # Alignment based on data type
        if df_results[col].dtype in ['float64', 'float32', 'int64', 'int32']:
            cell.alignment = number_alignment
        else:
            cell.alignment = cell_alignment

# Set row height for header
ws_results.row_dimensions[1].height = 35

print("    Applied scenario-based formatting to Results sheet")

# Format Column_Descriptions sheet
print("\n  Formatting 'Column_Descriptions' sheet...")
ws_desc = wb['Column_Descriptions']

# Apply header formatting
for cell in ws_desc[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Freeze panes
ws_desc.freeze_panes = 'A2'

# Set column widths
ws_desc.column_dimensions['A'].width = 35  # Column name
ws_desc.column_dimensions['B'].width = 70  # Description
ws_desc.column_dimensions['C'].width = 60  # Input data source

# Apply borders and alignment
for row_idx in range(2, len(df_descriptions) + 2):
    # Column A - Column name
    ws_desc[f'A{row_idx}'].border = border
    ws_desc[f'A{row_idx}'].alignment = cell_alignment
    ws_desc[f'A{row_idx}'].font = Font(name='Calibri', size=10, bold=True)

    # Column B - Description
    ws_desc[f'B{row_idx}'].border = border
    ws_desc[f'B{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_desc[f'B{row_idx}'].font = Font(name='Calibri', size=10)

    # Column C - Input data source
    ws_desc[f'C{row_idx}'].border = border
    ws_desc[f'C{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_desc[f'C{row_idx}'].font = Font(name='Calibri', size=10, italic=True)
    ws_desc[f'C{row_idx}'].fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')  # Light yellow

# Set row height for header
ws_desc.row_dimensions[1].height = 35

print("    Applied formatting to Column_Descriptions sheet")

# Save formatted workbook
wb.save(output_file)
print(f"\n  Saved formatted workbook: {output_file}")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "="*80)
print("FORMATTING COMPLETE")
print("="*80)

print(f"\nFormatted Excel file:")
print(f"  File: {output_file}")
print(f"  Sheets: 2")
print(f"    - Results: {len(df_results)} rows x {len(df_results.columns)} columns")
print(f"    - Column_Descriptions: {len(df_descriptions)} descriptions")

print("\nNew Features in This Version:")
print("  - Inundation comparison columns (3 new columns)")
print("  - Scenario-based row coloring (light blue vs light gray)")
print("  - Thicker borders between scenario groups")
print("  - Frozen columns (reach_id + Scenario)")
print("  - Third column showing input data sources")
print("  - Logical column ordering for reviewers")
print("  - Reaches sorted by HYDSEQ (upstream to downstream)")

print("\nFormatting Applied:")
print("  - Numbers rounded (Loads: 5 decimals, Percentages: 2 decimals)")
print("  - Professional header styling")
print("  - Optimized column widths")
print("  - Cell borders")
print("  - Proper alignment")

print("\n" + "="*80)
